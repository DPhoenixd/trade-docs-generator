from __future__ import annotations

import html
import re
import ctypes
from ctypes import wintypes
from dataclasses import replace
from datetime import date
from pathlib import Path

import pandas as pd
import streamlit as st

from trade_docs.calculations import (
    INVOICE_COLUMNS,
    ROLL_COLUMNS,
    apply_gross_weight,
    calculate_invoice_lines,
    calculate_rolls,
    empty_invoice_dataframe,
    empty_roll_dataframe,
    parse_roll_text,
    sample_roll_dataframe,
    validate_invoice_order,
    validate_order,
)
from trade_docs.exchange import fetch_usd_cny
from trade_docs.excel_writer import write_invoice_pair, write_packing_list
from trade_docs.fabric_db import discover_fabric_database, find_fabric, load_fabric_master, load_price_rules, save_fabric_record
from trade_docs.models import FabricRecord, OrderInfo
from trade_docs.numbering import suggest_doc_number
from trade_docs.ppo_parser import PPO_LINE_COLUMNS, empty_ppo_lines, parse_buyer_information, parse_ppo_pdf


BASE_DIR = Path.cwd()


def main() -> None:
    st.set_page_config(page_title="外贸单据工具", layout="wide", initial_sidebar_state="collapsed")
    _style()
    _init_state()

    with st.expander("模板和数据库", expanded=False):
        paths = _source_files()

    fabric_df, price_rules, db_path = _load_data(paths)

    if not st.session_state.get("module"):
        _home(db_path, fabric_df, price_rules)
        return

    module = st.session_state.module
    _module_header(module)

    if module == "invoice":
        _invoice_pdf_import(paths)

    buyer_template = paths["ci_template"] if module == "invoice" else paths["pl_template"]
    order, fabric, exchange_rate, exchange_source = _order_inputs(
        paths=paths,
        fabric_df=fabric_df,
        price_rules=price_rules,
        buyer_template=buyer_template,
        module=module,
    )
    if module == "invoice":
        default_price = float(st.session_state.get("invoice_default_price_usd", 18.2) or 18.2)
        invoice_input, quantity_unit = _invoice_lines_inputs(default_price, fabric)
        invoice_lines = _compute_invoice(fabric, invoice_input)
        order = replace(order, advance_payment_usd=round(float(invoice_lines["amount_usd"].sum()) * 0.3, 2) if not invoice_lines.empty else 0.0)
        errors, warnings = _validate_invoice(fabric, order, invoice_input, quantity_unit)
        if not errors and invoice_lines.empty:
            errors.append("P.I/C.I 明细没有可计算行，请检查 Quantity、Unit 和 USD/KG。")
        _summary_strip(order, fabric, invoice_lines, errors, mode="invoice")
        quantity_unit = _invoice_output_unit(invoice_input)
        _invoice_generate(paths, order, fabric, invoice_lines, quantity_unit, errors, warnings)
    else:
        default_usd_price = float(st.session_state.get("packing_default_usd_price", 10.3) or 10.3)
        rolls_input = _rolls_inputs(order.art_no, default_usd_price)
        computed_rolls, summary = _compute_packing(fabric, rolls_input)
        errors, warnings = _validate_packing(fabric, order, computed_rolls, summary)
        _summary_strip(order, fabric, summary, errors, mode="packing")
        _packing_generate(paths, order, fabric, computed_rolls, summary, errors, warnings)


def _init_state() -> None:
    if "pi_no" not in st.session_state:
        st.session_state.pi_no = suggest_doc_number(today=date.today())
    if "ci_no" not in st.session_state:
        st.session_state.ci_no = st.session_state.pi_no
    if "exchange_rate" not in st.session_state:
        st.session_state.exchange_rate = 7.20
        st.session_state.exchange_source = "手动默认值"
    if "rolls_df" not in st.session_state:
        st.session_state.rolls_df = empty_roll_dataframe()
    if "invoice_df" not in st.session_state:
        st.session_state.invoice_df = empty_invoice_dataframe()
    if "invoice_quantity_unit" not in st.session_state:
        st.session_state.invoice_quantity_unit = "Yard"
    if "ppo_summary" not in st.session_state:
        st.session_state.ppo_summary = {}
    if "ppo_lines_df" not in st.session_state:
        st.session_state.ppo_lines_df = empty_ppo_lines()
    if "buyer_text" not in st.session_state:
        st.session_state.buyer_text = ""
    if "invoice_default_price_usd" not in st.session_state:
        st.session_state.invoice_default_price_usd = 18.2
    if "fabric_intake_text" not in st.session_state:
        st.session_state.fabric_intake_text = ""
    if "fabric_ocr_text" not in st.session_state:
        st.session_state.fabric_ocr_text = ""
    if "roll_ocr_text" not in st.session_state:
        st.session_state.roll_ocr_text = ""
    if "packing_default_usd_price" not in st.session_state:
        st.session_state.packing_default_usd_price = 10.3


def _home(db_path: Path, fabric_df: pd.DataFrame, price_rules: pd.DataFrame) -> None:
    st.markdown(
        """
        <div class="hero">
          <h1>外贸单据工具</h1>
          <p>先选择要生成的单据。进去后只显示那一种单据需要的内容。</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    left, right = st.columns(2, gap="large")
    with left:
        st.markdown(
            """
            <div class="entry-card">
              <h2>P.I / C.I</h2>
              <p>生成 Proforma Invoice 和 Commercial Invoice。</p>
              <span>发票入口</span>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button("进入 P.I / C.I", type="primary", width="stretch"):
            st.session_state.module = "invoice"
            st.rerun()
    with right:
        st.markdown(
            """
            <div class="entry-card">
              <h2>Packing List</h2>
              <p>生成装箱单，按颜色 / LOT 展开每条布。</p>
              <span>装箱单入口</span>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button("进入 Packing List", width="stretch"):
            st.session_state.module = "packing"
            st.rerun()

    st.markdown(
        f"""
        <div class="tiny-status">
          <div><span>面料数据库</span><b>{html.escape(db_path.name)}</b></div>
          <div><span>面料记录</span><b>{len(fabric_df)}</b></div>
          <div><span>价格规则</span><b>{0 if price_rules.empty else len(price_rules)}</b></div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _module_header(module: str) -> None:
    title = "P.I / C.I 发票" if module == "invoice" else "Packing List 装箱单"
    desc = "生成两份 Excel：Proforma Invoice + Commercial Invoice。" if module == "invoice" else "生成一份 Excel：Packing List。"
    col1, col2 = st.columns([1, 0.18])
    with col1:
        st.markdown(f"<div class='module-title'><h1>{title}</h1><p>{desc}</p></div>", unsafe_allow_html=True)
    with col2:
        if st.button("换入口", width="stretch"):
            st.session_state.module = None
            st.rerun()


def _source_files() -> dict[str, Path | None]:
    cols = st.columns(5)
    with cols[0]:
        ppo_pdf = _file_select("PPO PDF", [".pdf"], "POUT26VE0011181A.pdf")
    with cols[1]:
        pl_template = _file_select("Packing List 模板", [".xlsx"], "Packing List-POUT25VE0011181A-25A109A.xlsx")
    with cols[2]:
        ci_template = _file_select("PI / CI 模板", [".xlsx"], "CI-POUT25VE0011181A-25A109A.xlsx")
    with cols[3]:
        fabric_db = _file_select("面料数据库", [".json", ".csv", ".xlsx"], "fabric_master_en.csv")
    with cols[4]:
        price_rules = _file_select("价格规则", [".csv", ".xlsx"], "fabric_price_rules.csv")

    with st.expander("上传替换", expanded=False):
        upload_cols = st.columns(5)
        with upload_cols[0]:
            ppo_pdf = _uploaded_or_current("PPO PDF", ["pdf"], ppo_pdf)
        with upload_cols[1]:
            pl_template = _uploaded_or_current("Packing List", ["xlsx"], pl_template)
        with upload_cols[2]:
            ci_template = _uploaded_or_current("PI / CI", ["xlsx"], ci_template)
        with upload_cols[3]:
            fabric_db = _uploaded_or_current("面料数据库", ["json", "csv", "xlsx"], fabric_db)
        with upload_cols[4]:
            price_rules = _uploaded_or_current("价格规则", ["csv", "xlsx"], price_rules)
    return {
        "ppo_pdf": ppo_pdf,
        "pl_template": pl_template,
        "ci_template": ci_template,
        "fabric_db": fabric_db,
        "price_rules": price_rules,
    }


@st.cache_data(show_spinner=False)
def _load_master_cached(path: str) -> pd.DataFrame:
    return load_fabric_master(Path(path))


@st.cache_data(show_spinner=False)
def _load_rules_cached(path: str) -> pd.DataFrame:
    return load_price_rules(Path(path))


def _load_data(paths: dict[str, Path | None]) -> tuple[pd.DataFrame, pd.DataFrame, Path]:
    db_path = paths.get("fabric_db") or discover_fabric_database(BASE_DIR)
    if not db_path:
        st.error("没有找到 fabric_database_en.json / fabric_master_en.csv / fabric_database_en.xlsx")
        st.stop()
    try:
        fabric_df = _load_master_cached(str(db_path))
    except Exception as exc:  # noqa: BLE001
        st.error(f"读取面料数据库失败：{exc}")
        st.stop()

    rules_path = paths.get("price_rules")
    price_rules = _load_rules_cached(str(rules_path)) if rules_path else pd.DataFrame()
    return fabric_df, price_rules, Path(db_path)


def _invoice_pdf_import(paths: dict[str, Path | None]) -> pd.DataFrame:
    st.markdown("<div class='section-label'>1. 客户订单 PDF</div>", unsafe_allow_html=True)
    pdf_path = paths.get("ppo_pdf")
    upload_cols = st.columns([0.5, 0.2, 0.3])
    with upload_cols[0]:
        uploaded_pdf = st.file_uploader("拖入 / 选择本次客户 PPO / Global PDF", type=["pdf"], key="invoice_ppo_pdf_upload")
    if uploaded_pdf is not None:
        upload_dir = BASE_DIR / ".streamlit_uploads"
        upload_dir.mkdir(exist_ok=True)
        pdf_path = upload_dir / uploaded_pdf.name
        pdf_path.write_bytes(uploaded_pdf.getbuffer())
        paths["ppo_pdf"] = pdf_path
        upload_cols[2].success(f"已选择：{uploaded_pdf.name}")
    elif upload_cols[1].button("读取剪贴板文件", width="stretch"):
        clip_path, message = _read_clipboard_pdf()
        if clip_path:
            pdf_path = clip_path
            paths["ppo_pdf"] = pdf_path
            upload_cols[2].success(f"已读取：{pdf_path.name}")
        else:
            upload_cols[2].warning(message)
    elif pdf_path:
        upload_cols[2].caption(f"当前文件：{pdf_path.name}")
    else:
        upload_cols[2].warning("请先拖入、选择或复制客户 PDF")

    with st.expander("更多导入方式", expanded=False):
        path_text = st.text_input("粘贴本地 PDF 路径", value="", placeholder=r"D:\...\POUT26VE0011181A.pdf")
        if st.button("使用这个 PDF 路径", disabled=not path_text.strip(), width="stretch"):
            candidate = Path(path_text.strip().strip('"'))
            if candidate.exists() and candidate.suffix.lower() == ".pdf":
                paths["ppo_pdf"] = candidate
                pdf_path = candidate
                st.success(f"已选择：{candidate.name}")
            else:
                st.error("这个路径不是有效 PDF 文件")

    cols = st.columns([0.22, 0.22, 0.56])
    parse_disabled = pdf_path is None
    if cols[0].button("解析客户订单", type="primary", disabled=parse_disabled, width="stretch"):
        try:
            summary, lines = parse_ppo_pdf(pdf_path)
        except Exception as exc:  # noqa: BLE001
            st.error(f"解析 PPO PDF 失败：{exc}")
        else:
            st.session_state.ppo_summary = summary
            st.session_state.ppo_lines_df = lines if not lines.empty else empty_ppo_lines()
            if summary.get("buyer_info_raw"):
                st.session_state.buyer_text = summary["buyer_info_raw"]
            st.success("已解析客户订单，下面可以审阅识别结果")
    if cols[1].button("清空订单识别", width="stretch"):
        st.session_state.ppo_summary = {}
        st.session_state.ppo_lines_df = empty_ppo_lines()

    cols[2].caption("支持拖拽 PDF 到上传框，或在资源管理器复制 PDF 后读取剪贴板文件。")

    summary = st.session_state.get("ppo_summary") or {}
    lines_df = _ensure_ppo_lines_columns(st.session_state.get("ppo_lines_df"))
    _ppo_summary_box(summary, lines_df)
    edited = st.data_editor(
        lines_df,
        num_rows="dynamic",
        width="stretch",
        height=220,
        column_config={
            "use": st.column_config.CheckboxColumn("导入", default=True),
            "gmt_color_code": st.column_config.TextColumn("GMT COLOR"),
            "fabric_combo_name": st.column_config.TextColumn("COLOR"),
            "source_fabric_code": st.column_config.TextColumn("PDF Fabric Code"),
            "fabric_total_yards": st.column_config.NumberColumn("Fabric Total(Yds)", min_value=0.0, step=1.0, format="%.2f"),
            "ppo_pur_qty_yards": st.column_config.NumberColumn("PPO PUR QTY(YDS) 参考", min_value=0.0, step=1.0, format="%.2f"),
            "ppo_qty_yards": st.column_config.NumberColumn("PPO QTY(Yds)", min_value=0.0, step=1.0, format="%.2f"),
            "source_component": st.column_config.TextColumn("来源"),
        },
        hide_index=True,
    )
    st.session_state.ppo_lines_df = _ensure_ppo_lines_columns(pd.DataFrame(edited))
    return st.session_state.ppo_lines_df


def _ppo_summary_box(summary: dict, lines_df: pd.DataFrame) -> None:
    if not summary and lines_df.empty:
        st.info("先解析客户 PPO PDF。识别结果会暂存在这里，生成前都可以修改。")
        return
    total_yards = float(pd.to_numeric(lines_df.get("fabric_total_yards"), errors="coerce").fillna(0).sum()) if not lines_df.empty else 0.0
    buyer = summary.get("buyer") or summary.get("customer") or "-"
    st.markdown(
        f"""
        <div class="import-box">
          <div><span>PPO</span><b>{html.escape(summary.get("ppo_no") or "-")}</b></div>
          <div><span>Style</span><b>{html.escape(summary.get("style_no") or "-")}</b></div>
          <div><span>Buyer</span><b>{html.escape(buyer)}</b></div>
          <div><span>颜色行</span><b>{len(lines_df)}</b></div>
          <div><span>Fabric Total</span><b>{total_yards:,.2f} YDS</b></div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _ensure_ppo_lines_columns(df: pd.DataFrame | None) -> pd.DataFrame:
    if df is None or df.empty:
        return empty_ppo_lines()
    out = df.copy()
    for column in PPO_LINE_COLUMNS:
        if column not in out.columns:
            out[column] = True if column == "use" else None
    return out[PPO_LINE_COLUMNS]


def _order_inputs(
    paths: dict[str, Path | None],
    fabric_df: pd.DataFrame,
    price_rules: pd.DataFrame,
    buyer_template: Path | None,
    module: str,
) -> tuple[OrderInfo, FabricRecord | None, float | None, str]:
    if module == "invoice":
        return _invoice_order_inputs(paths, fabric_df, price_rules, buyer_template)

    return _packing_order_inputs(paths, fabric_df, price_rules, buyer_template)


def _packing_order_inputs(
    paths: dict[str, Path | None],
    fabric_df: pd.DataFrame,
    price_rules: pd.DataFrame,
    buyer_template: Path | None,
) -> tuple[OrderInfo, FabricRecord | None, float | None, str]:
    st.markdown("<div class='section-label'>1. 面料属性和装箱单抬头</div>", unsafe_allow_html=True)
    left, right = st.columns([1.05, 0.95], gap="large")

    with left:
        st.caption("Packing List 的面料属性通常沿用 P.I/C.I：面料编号、成分、克重、幅宽、量化、纸筒空差。")
        fabric_codes = sorted([c for c in fabric_df["fabric_code"].dropna().astype(str).unique() if c.strip()])
        default_code = "6373D" if "6373D" in fabric_codes else (fabric_codes[0] if fabric_codes else "")
        fabric_mode = st.radio(
            "面料来源",
            ["使用面料数据库预设", "手动输入面料参数"],
            horizontal=True,
            key="packing_fabric_mode",
        )
        manual_fabric = fabric_mode == "手动输入面料参数"
        if manual_fabric:
            if "packing_manual_fabric_code" not in st.session_state:
                st.session_state.packing_manual_fabric_code = ""
            fabric_code = st.text_input(
                "面料编号 / 款号",
                key="packing_manual_fabric_code",
                placeholder="例如 25A109A 或 6373D",
            ).strip()
            grade = st.text_input("质量等级", value="First Grade", key="packing_manual_grade")
            base_fabric = None
            fabric_key = _fabric_widget_prefix(f"packing_{fabric_code}", grade)
            fabric = _fabric_edit_panel(base_fabric, fabric_code, grade, fabric_key) if fabric_code else None
            if not fabric_code:
                st.warning("手动模式下请填写面料编号，并在“面料数据人工修正”里补齐量化和纸筒空差。")
        else:
            fabric_code = st.selectbox(
                "面料数据预设",
                fabric_codes,
                index=fabric_codes.index(default_code) if default_code in fabric_codes else 0,
                key="packing_fabric_preset_choice",
            )
            grade_options = _grade_options(fabric_df, fabric_code)
            grade = st.selectbox("质量等级", grade_options, index=0, key="packing_grade")
            base_fabric = find_fabric(fabric_df, fabric_code, grade)
            fabric_key = _fabric_widget_prefix(f"packing_{fabric_code}", grade)
            fabric = _fabric_edit_panel(base_fabric, fabric_code, grade, fabric_key)

        if fabric:
            _fabric_line(fabric, "Packing List 使用面料量化和纸筒空差计算 Meter / Yard / Gross Weight")
        else:
            st.error("没有可用面料数据。请选择数据库预设，或手动输入面料参数。")
        _price_rules(fabric_code, price_rules, selected_grade=grade)
        _fabric_database_intake(paths, fabric, fabric_key)

    with right:
        row1 = st.columns([1, 0.85])
        po_no = row1[0].text_input("PO NO", value="POUT26VE0011181A", key="packing_po_no")
        art_no = row1[1].text_input("ITEM / ART NO", value="25A109A", key="packing_art_no")
        buyer = st.text_input("TO / Buyer", value=_default_buyer(buyer_template, "packing") or "COTTEX LLP", key="packing_buyer")
        row2 = st.columns([1, 0.8])
        pi_no = row2[0].text_input("P/I NO", value=st.session_state.pi_no, key="packing_pi_no")
        order_date = row2[1].date_input("日期", value=date.today(), key="packing_date").isoformat()
        st.session_state.packing_default_usd_price = st.number_input(
            "默认 USD/KG（来自 P.I/C.I）",
            min_value=0.0,
            value=float(st.session_state.packing_default_usd_price),
            step=0.1,
            format="%.4f",
            key="packing_default_price_input",
        )
        advance_payment = st.number_input("Advance payment USD（如需右侧汇总显示）", min_value=0.0, value=0.0, step=100.0, format="%.2f")
        note = st.text_input("NOTE", value="")
        output_dir = Path(st.text_input("保存目录", value=str(BASE_DIR / "outputs")))

    order = OrderInfo(
        po_no=po_no,
        art_no=art_no,
        fabric_code=fabric.fabric_code if fabric else fabric_code,
        buyer=buyer,
        pi_no=pi_no,
        ci_no=pi_no,
        order_date=order_date,
        advance_payment_usd=float(advance_payment),
        output_dir=output_dir,
        note=note,
    )
    return order, fabric, None, "Packing List 使用 USD/KG，不需要汇率"


def _invoice_order_inputs(
    paths: dict[str, Path | None],
    fabric_df: pd.DataFrame,
    price_rules: pd.DataFrame,
    buyer_template: Path | None,
) -> tuple[OrderInfo, FabricRecord | None, float | None, str]:
    st.markdown("<div class='section-label'>2. 面料预设和客户档案</div>", unsafe_allow_html=True)
    summary = st.session_state.get("ppo_summary") or {}
    left, right = st.columns([1.05, 0.95], gap="large")

    with left:
        fabric_codes = sorted([c for c in fabric_df["fabric_code"].dropna().astype(str).unique() if c.strip()])
        default_code = _best_fabric_default(fabric_codes)
        fabric_mode = st.radio(
            "面料来源",
            ["使用面料数据库预设", "手动输入面料参数"],
            horizontal=True,
            key="invoice_fabric_mode",
        )
        manual_fabric = fabric_mode == "手动输入面料参数"
        if manual_fabric:
            if "invoice_manual_fabric_code" not in st.session_state:
                st.session_state.invoice_manual_fabric_code = summary.get("quality_code") or ""
            fabric_code = st.text_input(
                "面料编号",
                key="invoice_manual_fabric_code",
                placeholder="例如 25A109A。这里可以 Ctrl+A 后直接删除。",
            )
            fabric_code = str(fabric_code or "").strip()
            grade = st.text_input("质量等级", value="First Grade")
            base_fabric = None
            fabric_key = _fabric_widget_prefix(fabric_code, grade)
            if fabric_code:
                fabric = _fabric_edit_panel(base_fabric, fabric_code, grade, fabric_key)
            else:
                fabric = None
                st.warning("手动模式下请填写面料编号，或切回数据库预设。")
        else:
            fabric_choice = st.selectbox(
                "面料数据预设",
                fabric_codes,
                index=fabric_codes.index(default_code) if default_code in fabric_codes else 0,
                key="invoice_fabric_preset_choice",
            )
            fabric_code = fabric_choice
            grade_options = _grade_options(fabric_df, fabric_code)
            grade = st.selectbox("质量等级", grade_options, index=0)
            base_fabric = find_fabric(fabric_df, fabric_code, grade)
            fabric_key = _fabric_widget_prefix(fabric_code, grade)
            fabric = _fabric_edit_panel(base_fabric, fabric_code, grade, fabric_key)
        if fabric and not manual_fabric:
            _fabric_line(fabric, "P.I/C.I 使用 USD 销售单价")
        elif fabric:
            st.info("当前使用手动面料数据生成本次单据。点“写入面料数据库”后才会保存到数据库。")
            _fabric_line(fabric, "P.I/C.I 使用 USD 销售单价")
        else:
            st.error("面料编号未在数据库中找到，请手动补充或检查编号")
        price_hint = _price_hint(fabric_code, grade, price_rules)
        st.session_state.invoice_default_price_usd = st.number_input(
            "销售 USD/KG 预设",
            min_value=0.0,
            value=float(st.session_state.invoice_default_price_usd),
            step=0.1,
            format="%.4f",
            help="导入当前明细时会自动填入每个颜色，之后仍可在审阅表里修正。",
        )
        if price_hint:
            st.caption(price_hint)
        _price_rules(fabric_code, price_rules, selected_grade=grade)
        _fabric_database_intake(paths, fabric, fabric_key)

    with right:
        default_buyer = summary.get("buyer") or _default_buyer(buyer_template, "invoice")
        if not st.session_state.buyer_text and summary.get("buyer_info_raw"):
            st.session_state.buyer_text = summary["buyer_info_raw"]
        st.markdown("<div class='sub-label'>客户档案（Buyer’s information）</div>", unsafe_allow_html=True)
        buyer_text = st.text_area(
            "复制/粘贴客户信息",
            value=st.session_state.buyer_text or default_buyer,
            height=120,
            placeholder="第一行客户名称，后面几行地址、电话、传真等",
        )
        st.session_state.buyer_text = buyer_text
        if st.button("识别 Buyer 信息", width="stretch"):
            parsed_buyer = parse_buyer_information(buyer_text)
            st.session_state.ppo_summary = {**summary, **parsed_buyer}
            st.rerun()

        parsed_buyer = parse_buyer_information(st.session_state.buyer_text)
        buyer = (summary.get("buyer") or parsed_buyer.get("buyer") or default_buyer or "").strip()
        buyer_address = (summary.get("buyer_address") or parsed_buyer.get("buyer_address") or "").strip()
        po_no = (summary.get("ppo_no") or "POUT26VE0011181A").strip()
        art_no = (summary.get("style_no") or "25A109A").strip()
        with st.expander("订单字段人工修正", expanded=False):
            fix = st.columns(3)
            po_no = fix[0].text_input("PO NO", value=po_no)
            art_no = fix[1].text_input("ART NO", value=art_no)
            buyer = fix[2].text_input("Buyer", value=buyer)
            buyer_address = st.text_area("Buyer 地址", value=buyer_address, height=72)

        st.markdown(
            f"""
            <div class="doc-box">
              <div><span>P/I NO</span><b>{html.escape(st.session_state.pi_no)}</b></div>
              <div><span>C/I NO</span><b>{html.escape(st.session_state.ci_no)}</b></div>
              <div><span>预付款</span><b>Total Amount × 30%</b></div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    order = OrderInfo(
        po_no=po_no,
        art_no=art_no,
        fabric_code=fabric_code,
        buyer=buyer,
        pi_no=st.session_state.pi_no,
        ci_no=st.session_state.ci_no,
        order_date=date.today().isoformat(),
        advance_payment_usd=0.0,
        output_dir=BASE_DIR / "outputs",
        buyer_address=buyer_address,
    )
    return order, fabric, None, "P.I/C.I 使用 USD 销售单价"


def _fabric_edit_panel(base: FabricRecord | None, fabric_code: str, grade: str, key_prefix: str) -> FabricRecord | None:
    values = _fabric_values(base, fabric_code, grade)
    with st.expander("面料数据人工修正", expanded=False):
        st.caption("这里的修改会立刻用于当前 P.I/C.I 计算和 Excel 生成；不会自动写入数据库。")
        row1 = st.columns([0.7, 1, 1, 0.8])
        edited_code = row1[0].text_input("面料编号", value=values["fabric_code"], key=f"{key_prefix}_code")
        fabric_name_en = row1[1].text_input("英文品名", value=values["fabric_name_en"], key=f"{key_prefix}_name_en")
        fabric_name_cn = row1[2].text_input("中文品名", value=values["fabric_name_cn"], key=f"{key_prefix}_name_cn")
        quality_grade_en = row1[3].text_input("质量等级 EN", value=values["quality_grade_en"], key=f"{key_prefix}_grade_en")

        composition_en = st.text_area("英文成分", value=values["composition_en"], height=72, key=f"{key_prefix}_comp_en")
        row2 = st.columns(4)
        weight_en = row2[0].text_input("克重", value=values["weight_en"], key=f"{key_prefix}_weight_en")
        width_en = row2[1].text_input("幅宽", value=values["width_en"], key=f"{key_prefix}_width_en")
        quantification = row2[2].number_input(
            "量化 m/kg",
            min_value=0.0,
            value=float(values["quantification_m_per_kg"] or 0.0),
            step=0.01,
            format="%.4f",
            key=f"{key_prefix}_quantification",
        )
        tube_allowance = row2[3].number_input(
            "纸筒空差 kg/roll",
            min_value=0.0,
            value=float(values["tube_plus_allowance_kg_per_roll"] or 0.0),
            step=0.1,
            format="%.2f",
            key=f"{key_prefix}_tube",
        )
        remarks_en = st.text_input("备注 EN", value=values["remarks_en"], key=f"{key_prefix}_remarks_en")

    if not str(edited_code).strip():
        return None
    raw = dict(base.raw) if base else {}
    return FabricRecord(
        fabric_code=str(edited_code).strip(),
        quality_grade_cn=base.quality_grade_cn if base else "",
        quality_grade_en=quality_grade_en.strip() or str(grade).strip(),
        fabric_name_cn=fabric_name_cn.strip(),
        fabric_name_en=fabric_name_en.strip(),
        composition_cn=base.composition_cn if base else "",
        composition_en=composition_en.strip(),
        width_cn=base.width_cn if base else "",
        width_en=width_en.strip(),
        weight_cn=base.weight_cn if base else "",
        weight_en=weight_en.strip(),
        quantification_m_per_kg=float(quantification) if quantification else None,
        tube_plus_allowance_kg_per_roll=float(tube_allowance) if tube_allowance or tube_allowance == 0 else None,
        reference_roll_weight_kg=base.reference_roll_weight_kg if base else None,
        remarks_cn=base.remarks_cn if base else "",
        remarks_en=remarks_en.strip(),
        raw=raw,
    )


def _fabric_database_intake(paths: dict[str, Path | None], fabric: FabricRecord | None, key_prefix: str) -> None:
    db_path = paths.get("fabric_db")
    if st.session_state.get("fabric_db_write_message"):
        st.success(st.session_state.pop("fabric_db_write_message"))
    with st.expander("反向录入面料数据库", expanded=False):
        st.caption("截图后直接读取剪贴板；也支持粘贴文字。上传文件只是备用方式。写入前会自动备份原 CSV。")
        clip_cols = st.columns([0.32, 0.68])
        if clip_cols[0].button("读取剪贴板截图", key="fabric_read_clipboard_image", width="stretch"):
            text, message = _ocr_clipboard_image()
            if text:
                st.session_state.fabric_ocr_text = text
                st.session_state.fabric_intake_text = text
                st.success(message)
                st.rerun()
            else:
                st.warning(message)
        clip_cols[1].caption("先用 Win+Shift+S 或微信/QQ截图复制到剪贴板，然后点左侧按钮。")

        uploaded = st.file_uploader("备用：上传截图文件", type=["png", "jpg", "jpeg", "webp"], key="fabric_ocr_upload")
        ocr_cols = st.columns([0.24, 0.76])
        if ocr_cols[0].button("识别上传截图", key="fabric_ocr_uploaded_image", disabled=uploaded is None, width="stretch"):
            text, message = _ocr_fabric_image(uploaded)
            if text:
                st.session_state.fabric_ocr_text = text
                st.session_state.fabric_intake_text = text
                st.success(message)
                st.rerun()
            else:
                st.warning(message)
        ocr_cols[1].caption("OCR 依赖本机 Tesseract；如果识别失败，把截图里的文字复制到下面文本框也可以。")

        intake_text = st.text_area(
            "手动输入 / 复制粘贴面料信息",
            value=st.session_state.fabric_intake_text,
            height=130,
            placeholder="例：面料编号 6373D；英文品名 Air Wool Fleece；成分 ...；克重 380 gsm；幅宽 170 cm；量化 1.55；纸筒空差 1.0",
        )
        st.session_state.fabric_intake_text = intake_text
        candidate = fabric
        if intake_text.strip():
            parsed = _parse_fabric_text(intake_text, fabric)
            candidate = _fabric_candidate_editor(parsed, intake_text)
        action_cols = st.columns([0.28, 0.28, 0.44])
        if action_cols[0].button("套用识别结果", disabled=candidate is None, width="stretch"):
            _apply_fabric_to_editor(candidate, key_prefix)
            st.rerun()
        if action_cols[1].button("写入面料数据库", type="primary", disabled=candidate is None, width="stretch"):
            errors = _validate_fabric_record(candidate)
            if errors:
                for error in errors:
                    st.error(error)
            elif db_path is None:
                st.error("没有选择面料数据库")
            else:
                try:
                    action, backup_path = save_fabric_record(db_path, candidate)
                except Exception as exc:  # noqa: BLE001
                    st.error(f"写入失败：{exc}")
                else:
                    _load_master_cached.clear()
                    st.session_state.fabric_db_write_message = (
                        f"{'已更新' if action == 'updated' else '已新增'}面料 {candidate.fabric_code}，原数据库已备份：{backup_path.name}"
                    )
                    st.rerun()
        action_cols[2].caption("建议先套用并检查当前面料卡片，再写入数据库。")


def _fabric_values(base: FabricRecord | None, fabric_code: str, grade: str) -> dict[str, object]:
    return {
        "fabric_code": base.fabric_code if base else fabric_code,
        "quality_grade_en": base.quality_grade_en if base else grade,
        "fabric_name_cn": base.fabric_name_cn if base else "",
        "fabric_name_en": base.fabric_name_en if base else "",
        "composition_en": base.composition_en if base else "",
        "width_en": base.width_en if base else "",
        "weight_en": base.weight_en if base else "",
        "quantification_m_per_kg": base.quantification_m_per_kg if base else 0.0,
        "tube_plus_allowance_kg_per_roll": base.tube_plus_allowance_kg_per_roll if base else 0.0,
        "remarks_en": base.remarks_en if base else "",
    }


def _rolls_inputs(default_item: str, default_price_usd: float) -> pd.DataFrame:
    st.markdown("<div class='section-label'>2. 码单识别和细码审查</div>", unsafe_allow_html=True)
    st.caption("Packing List 的核心数据来自码单：每条布的 Net Weight/KG、颜色、LOT。截图可以复制粘贴读取，也可以拖入图片识别；识别结果必须在下方表格人工确认。")
    toolbar = st.columns([0.16, 0.16, 0.68])
    if toolbar[0].button("载入示例", width="stretch"):
        st.session_state.rolls_df = sample_roll_dataframe()
    if toolbar[1].button("清空", width="stretch"):
        st.session_state.rolls_df = empty_roll_dataframe()

    with st.expander("截图识别 / 文本粘贴", expanded=True):
        ocr_actions = st.columns([0.24, 0.34, 0.42])
        if ocr_actions[0].button("读取剪贴板截图", key="roll_read_clipboard_image", width="stretch"):
            text, message = _ocr_clipboard_image(prefer_roll_parser=True, default_price=st.session_state.get("packing_default_usd_price"))
            if text:
                st.session_state.roll_ocr_text = text
                st.success(message)
            else:
                st.warning(message)
        uploaded_roll_images = ocr_actions[1].file_uploader(
            "拖入码单图片（可多选）",
            type=["png", "jpg", "jpeg", "webp"],
            key="roll_ocr_upload",
            accept_multiple_files=True,
            label_visibility="collapsed",
        )
        if ocr_actions[2].button("识别拖入图片", key="roll_ocr_uploaded_image", disabled=not uploaded_roll_images, width="stretch"):
            text, message = _ocr_uploaded_images(uploaded_roll_images, default_price=st.session_state.get("packing_default_usd_price"))
            if text:
                st.session_state.roll_ocr_text = text
                st.success(message)
            else:
                st.warning(message)

        pasted = st.text_area(
            "OCR 结果 / 手动粘贴码单文本",
            value=st.session_state.roll_ocr_text,
            height=140,
            placeholder="建议格式：ITEM, COLOR, LOT, 细码数量, USD/KG\n例如：25A109A, NAVY/#0, LDA2603-0032, 17.6,20.6,21.4, 10.30",
        )
        st.session_state.roll_ocr_text = pasted
        paste_cols = st.columns([0.22, 0.18, 0.60])
        default_price = paste_cols[0].number_input("默认 USD/KG", min_value=0.0, value=float(default_price_usd), step=0.1, format="%.4f")
        st.session_state.packing_default_usd_price = default_price
        if paste_cols[1].button("解析到明细表", width="stretch"):
            st.session_state.rolls_df = parse_roll_text(pasted, default_item=default_item, default_price=default_price)

    current_rolls = st.session_state.rolls_df.copy()
    if "rmb_price_per_kg" in current_rolls.columns and "usd_price_per_kg" not in current_rolls.columns:
        current_rolls = current_rolls.rename(columns={"rmb_price_per_kg": "usd_price_per_kg"})
    for column in ROLL_COLUMNS:
        if column not in current_rolls.columns:
            current_rolls[column] = None
    current_rolls = current_rolls[ROLL_COLUMNS]
    edited = st.data_editor(
        current_rolls,
        num_rows="dynamic",
        width="stretch",
        height=360,
        column_config={
            "item": st.column_config.TextColumn("ITEM"),
            "color": st.column_config.TextColumn("COLOR"),
            "lot": st.column_config.TextColumn("LOT"),
            "net_weight_kg": st.column_config.NumberColumn("Net KG", min_value=0.0, step=0.1, format="%.2f"),
            "usd_price_per_kg": st.column_config.NumberColumn("USD/KG", min_value=0.0, step=0.1, format="%.4f"),
        },
        hide_index=True,
    )
    st.session_state.rolls_df = pd.DataFrame(edited, columns=ROLL_COLUMNS)
    return st.session_state.rolls_df


def _invoice_lines_inputs(default_price_usd: float, fabric: FabricRecord | None) -> tuple[pd.DataFrame, str]:
    st.markdown("<div class='section-label'>3. P.I / C.I 明细审阅</div>", unsafe_allow_html=True)
    _invoice_fabric_context(fabric)
    st.info("PDF 识别到的数量和单位会直接进入明细。系统后台按单位换算：Yard → Meter → KG，或 Meter → KG，或 KG → Meter/Yard，再用 KG × USD/KG 计算金额。")
    st.caption("这里只需要确认 COLOR、Quantity、Unit、USD/KG。COLOR 是文本，可以填英文颜色、色号、英文+色号。")
    toolbar = st.columns([0.24, 0.16, 0.60])
    if toolbar[0].button("生成/刷新审阅明细", type="primary", width="stretch"):
        st.session_state.invoice_df = _invoice_rows_from_ppo(st.session_state.ppo_lines_df, default_price_usd)
    if toolbar[1].button("清空明细", width="stretch"):
        st.session_state.invoice_df = empty_invoice_dataframe()
    quantity_unit = "Yard"
    st.session_state.invoice_quantity_unit = quantity_unit

    editor_columns = ["color", "quantity_input", "input_unit", "usd_price_per_kg"]
    current = st.session_state.invoice_df.copy()
    for column in INVOICE_COLUMNS:
        if column not in current.columns:
            current[column] = None
    current["color"] = current["color"].fillna("").astype(str)
    current["source_fabric_code"] = current["source_fabric_code"].fillna("").astype(str)
    current["source_note"] = current["source_note"].fillna("").astype(str)
    current["input_unit"] = current["input_unit"].fillna("Yard").astype(str).replace({"YDS": "Yard", "Yards": "Yard", "Meters": "Meter"})
    current.loc[~current["input_unit"].isin(["KG", "Meter", "Yard"]), "input_unit"] = "Yard"
    with st.form("invoice_detail_form", clear_on_submit=False):
        edited = st.data_editor(
            current[editor_columns],
            num_rows="dynamic",
            width="stretch",
            height=240,
            column_config={
                "color": st.column_config.TextColumn("COLOR", help="颜色可以是英文、色号、英文+色号，不要求数字。"),
                "quantity_input": st.column_config.NumberColumn("Quantity", min_value=0.0, step=0.1, format="%.2f"),
                "input_unit": st.column_config.SelectboxColumn("Unit", options=["Yard", "Meter", "KG"], required=True),
                "usd_price_per_kg": st.column_config.NumberColumn("USD/KG", min_value=0.0, step=0.1, format="%.2f"),
            },
            hide_index=True,
        )
        saved = st.form_submit_button("保存明细并重新计算", type="primary", width="stretch")
    if saved:
        edited_df = pd.DataFrame(edited)
        updated = edited_df.copy()
        for hidden_column in ["ppo_reference_yards", "source_fabric_code", "source_note"]:
            values = current[hidden_column].reindex(updated.index) if hidden_column in current.columns else pd.Series(index=updated.index, dtype=object)
            updated[hidden_column] = values.values
        for column in INVOICE_COLUMNS:
            if column not in updated.columns:
                updated[column] = None
        updated["color"] = updated["color"].fillna("").astype(str).str.strip()
        updated["art_no"] = updated["color"].fillna("").astype(str).str.strip()
        updated["input_unit"] = updated["input_unit"].fillna("Yard").astype(str).replace({"YDS": "Yard", "Yards": "Yard", "Meters": "Meter"})
        updated.loc[~updated["input_unit"].isin(["KG", "Meter", "Yard"]), "input_unit"] = "Yard"
        updated["ppo_reference_yards"] = pd.to_numeric(updated["ppo_reference_yards"], errors="coerce")
        updated["quantity_input"] = pd.to_numeric(updated["quantity_input"], errors="coerce")
        updated["ppo_reference_yards"] = updated["ppo_reference_yards"].fillna(updated["quantity_input"])
        st.session_state.invoice_df = updated[INVOICE_COLUMNS]
        st.rerun()
    with st.expander("换算公式", expanded=False):
        st.code("Unit = Yard:  Meter = Quantity × 0.9144; KG = Meter ÷ 量化\nUnit = Meter: KG = Quantity ÷ 量化\nUnit = KG:    Meter = Quantity × 量化; Yard = Meter ÷ 0.9144\nAmount USD = KG × USD/KG")
    return st.session_state.invoice_df, quantity_unit


def _invoice_fabric_context(fabric: FabricRecord | None) -> None:
    if not fabric:
        st.warning("还没有可用面料数据。请先在第 2 步选择或录入面料。")
        return
    st.markdown(
        f"""
        <div class="doc-box">
          <div><span>当前用于生成的面料</span><b>{html.escape(fabric.fabric_code or "-")}</b></div>
          <div><span>量化</span><b>{_format_quantity_factor(fabric.quantification_m_per_kg)} m/kg</b></div>
          <div><span>克重 / 幅宽</span><b>{html.escape((fabric.weight_en or "-") + " / " + (fabric.width_en or "-"))}</b></div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    if not fabric.quantification_m_per_kg:
        st.error("当前面料缺少量化，无法把 Yard/Meter 反算 KG。")


def _art_no_from_color(color: str) -> str:
    text = str(color or "").strip()
    return text


def _format_quantity_factor(value: float | None) -> str:
    if not value:
        return "缺失"
    return f"{float(value):.4f}".rstrip("0").rstrip(".")


def _invoice_rows_from_ppo(ppo_lines: pd.DataFrame, default_price_usd: float) -> pd.DataFrame:
    lines = _ensure_ppo_lines_columns(ppo_lines)
    if lines.empty:
        return empty_invoice_dataframe()
    grouped: dict[tuple[str, str], dict[str, object]] = {}
    for _, row in lines.iterrows():
        if not bool(row.get("use", True)):
            continue
        yards = _first_number(row.get("fabric_total_yards"), row.get("ppo_qty_yards"), row.get("ppo_pur_qty_yards"))
        if yards is None:
            continue
        color = str(row.get("fabric_combo_name") or row.get("gmt_color_code") or "").strip()
        source_code = str(row.get("source_fabric_code") or "").strip()
        key = (color, source_code)
        if key not in grouped:
            grouped[key] = {"yards": 0.0, "gmt_colors": []}
        grouped[key]["yards"] = float(grouped[key]["yards"]) + yards
        gmt_color = str(row.get("gmt_color_code") or "").strip()
        if gmt_color and gmt_color not in grouped[key]["gmt_colors"]:
            grouped[key]["gmt_colors"].append(gmt_color)
    rows = []
    for (color, source_code), item in grouped.items():
        note = "PPO PDF"
        if item["gmt_colors"]:
            note = "PPO: " + "/".join(item["gmt_colors"])
        art_no = _art_no_from_color(color)
        rows.append([color, art_no, item["yards"], item["yards"], "Yard", default_price_usd, source_code, note])
    return pd.DataFrame(rows, columns=INVOICE_COLUMNS) if rows else empty_invoice_dataframe()


def _compute_packing(fabric: FabricRecord | None, rolls_input: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    if not fabric or not fabric.quantification_m_per_kg:
        return pd.DataFrame(), pd.DataFrame()
    computed_rolls, summary = calculate_rolls(rolls_input, float(fabric.quantification_m_per_kg))
    if fabric.tube_plus_allowance_kg_per_roll is not None:
        summary = apply_gross_weight(summary, float(fabric.tube_plus_allowance_kg_per_roll))
    return computed_rolls, summary


def _compute_invoice(fabric: FabricRecord | None, invoice_input: pd.DataFrame) -> pd.DataFrame:
    if not fabric or not fabric.quantification_m_per_kg:
        return pd.DataFrame()
    return calculate_invoice_lines(invoice_input, float(fabric.quantification_m_per_kg))


def _invoice_output_unit(invoice_input: pd.DataFrame) -> str:
    if invoice_input.empty or "input_unit" not in invoice_input.columns:
        return "Yard"
    units = (
        invoice_input["input_unit"]
        .fillna("Yard")
        .astype(str)
        .replace({"YDS": "Yard", "Yards": "Yard", "Meters": "Meter"})
    )
    units = units[units.isin(["KG", "Meter", "Yard"])]
    unique_units = list(dict.fromkeys(units.tolist()))
    return unique_units[0] if len(unique_units) == 1 else "Yard"


def _validate_packing(
    fabric: FabricRecord | None,
    order: OrderInfo,
    computed_rolls: pd.DataFrame,
    summary: pd.DataFrame,
) -> tuple[list[str], list[str]]:
    if not fabric:
        return ["面料编号未在数据库中找到，请手动补充或检查编号"], []
    if not fabric.quantification_m_per_kg or fabric.quantification_m_per_kg <= 0:
        return ["当前面料缺少量化 m/kg，无法把细码 KG 换算成 Meter/Yard。请在第 1 步“面料数据人工修正”里填写量化。"], []
    return validate_order(fabric, order, computed_rolls, summary)


def _validate_invoice(
    fabric: FabricRecord | None,
    order: OrderInfo,
    invoice_lines: pd.DataFrame,
    quantity_unit: str,
) -> tuple[list[str], list[str]]:
    return validate_invoice_order(fabric, order, invoice_lines, quantity_unit)


def _invoice_generate(
    paths: dict[str, Path | None],
    order: OrderInfo,
    fabric: FabricRecord | None,
    invoice_lines: pd.DataFrame,
    quantity_unit: str,
    errors: list[str],
    warnings: list[str],
) -> None:
    st.markdown("<div class='section-label'>4. 生成 P.I / C.I</div>", unsafe_allow_html=True)
    _checks(errors, warnings, "可以生成 P.I / C.I")
    if not invoice_lines.empty:
        st.caption(f"预付款自动按 Total Amount × 30% 计算：${order.advance_payment_usd:,.2f}")
    _invoice_preview(invoice_lines, quantity_unit, fabric)
    disabled = bool(errors) or fabric is None or paths.get("ci_template") is None
    if st.button("生成 P.I + C.I", type="primary", disabled=disabled, width="stretch"):
        result = write_invoice_pair(paths["ci_template"], order, fabric, invoice_lines, quantity_unit=quantity_unit)
        st.success("P.I / C.I 已生成")
        st.write(f"P.I：`{result.pi_path}`")
        st.write(f"C.I：`{result.ci_path}`")
        st.write(f"日志：`{result.log_path}`")
        c1, c2 = st.columns(2)
        with result.pi_path.open("rb") as f:
            c1.download_button("下载 P.I", f, file_name=result.pi_path.name, width="stretch")
        with result.ci_path.open("rb") as f:
            c2.download_button("下载 C.I", f, file_name=result.ci_path.name, width="stretch")


def _packing_generate(
    paths: dict[str, Path | None],
    order: OrderInfo,
    fabric: FabricRecord | None,
    computed_rolls: pd.DataFrame,
    summary: pd.DataFrame,
    errors: list[str],
    warnings: list[str],
) -> None:
    st.markdown("<div class='section-label'>3. 生成 Packing List</div>", unsafe_allow_html=True)
    _checks(errors, warnings, "可以生成 Packing List")
    _preview(summary, computed_rolls)
    disabled = bool(errors) or fabric is None or paths.get("pl_template") is None
    if st.button("生成 Packing List", type="primary", disabled=disabled, width="stretch"):
        result = write_packing_list(paths["pl_template"], order, fabric, computed_rolls, summary)
        st.success("Packing List 已生成")
        st.write(f"Excel：`{result.path}`")
        st.write(f"日志：`{result.log_path}`")
        with result.path.open("rb") as f:
            st.download_button("下载 Packing List", f, file_name=result.path.name, width="stretch")


def _checks(errors: list[str], warnings: list[str], success_text: str) -> None:
    if warnings:
        for warning in warnings:
            st.warning(warning)
    if errors:
        for error in errors:
            st.error(error)
    else:
        st.success(success_text)


def _invoice_preview(invoice_lines: pd.DataFrame, quantity_unit: str, fabric: FabricRecord | None) -> None:
    if invoice_lines.empty:
        st.info("保存明细并重新计算后显示自动换算结果")
        return
    quantity_column = f"Quantity({quantity_unit})"
    source_column = {"KG": "total_net_weight_kg", "Meter": "total_meter", "Yard": "total_yard"}[quantity_unit]
    show = pd.DataFrame(
        {
            "Fabric Code": fabric.fabric_code if fabric else "",
            "Composition": (fabric.composition_en or fabric.composition_cn) if fabric else "",
            "COLOR": invoice_lines["color"],
            quantity_column: invoice_lines[source_column],
            "Quantity(KG)": invoice_lines["total_net_weight_kg"],
            "Quantity(Meter)": invoice_lines["total_meter"],
            "Quantity(Yard)": invoice_lines["total_yard"],
            "USD/KG": invoice_lines["usd_price_per_kg"],
            "Amount": invoice_lines["amount_usd"],
        }
    )
    base_columns = ["Fabric Code", "Composition", "COLOR", quantity_column, "Quantity(KG)", "USD/KG", "Amount"]
    support_columns = [col for col in ["Quantity(Meter)", "Quantity(Yard)"] if col not in base_columns]
    show_conversions = st.checkbox("显示 Meter / Yard 换算辅助列", value=True, key="invoice_preview_show_conversions")
    keep = base_columns + (support_columns if show_conversions else [])
    show = show[keep]
    for col in [quantity_column, "Quantity(KG)", "Quantity(Meter)", "Quantity(Yard)", "USD/KG", "Amount"]:
        if col in show.columns:
            show[col] = pd.to_numeric(show[col], errors="coerce").round(4)
    st.dataframe(show, width="stretch", hide_index=True)


def _preview(summary: pd.DataFrame, computed_rolls: pd.DataFrame) -> None:
    if not summary.empty:
        show_summary = summary.copy()
        numeric_cols = show_summary.select_dtypes(include=["number"]).columns
        show_summary[numeric_cols] = show_summary[numeric_cols].round(4)
        st.dataframe(show_summary, width="stretch", hide_index=True)
    with st.expander("细码明细", expanded=False):
        if computed_rolls.empty:
            st.info("录入码单后显示明细")
        else:
            show_rolls = computed_rolls.copy()
            for col in ["meter", "yard", "usd_price_per_kg", "amount_usd"]:
                show_rolls[col] = show_rolls[col].round(4)
            st.dataframe(show_rolls, width="stretch", hide_index=True)


def _summary_strip(order: OrderInfo, fabric: FabricRecord | None, summary: pd.DataFrame, errors: list[str], mode: str) -> None:
    total_kg = float(summary["total_net_weight_kg"].sum()) if not summary.empty and "total_net_weight_kg" in summary else 0.0
    count_label = "颜色"
    count_value = len(summary) if mode == "invoice" else int(summary["rolls"].sum()) if not summary.empty and "rolls" in summary else 0
    if mode == "packing":
        count_label = "条数"
    amount = float(summary["amount_usd"].sum()) if not summary.empty and "amount_usd" in summary else 0.0
    status = "待修正" if errors else "可生成"
    status_class = "bad" if errors else "good"
    fabric_label = fabric.display_name if fabric else "未匹配"
    st.markdown(
        f"""
        <div class="summary-row">
          <div><span>PO</span><b>{html.escape(order.po_no or "-")}</b></div>
          <div><span>面料</span><b>{html.escape(fabric_label)}</b></div>
          <div><span>净重</span><b>{total_kg:,.2f} KG</b></div>
          <div><span>{count_label}</span><b>{count_value}</b></div>
          <div><span>金额</span><b>${amount:,.2f}</b></div>
          <div><span>状态</span><b class="{status_class}">{status}</b></div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _fabric_line(fabric: FabricRecord, exchange_source: str) -> None:
    st.markdown(
        f"""
        <div class="fabric-line">
          <div><span>品名</span><b>{html.escape(fabric.fabric_name_en or fabric.fabric_name_cn or "-")}</b></div>
          <div><span>克重</span><b>{html.escape(fabric.weight_en or fabric.weight_cn or "-")}</b></div>
          <div><span>幅宽</span><b>{html.escape(fabric.width_en or fabric.width_cn or "-")}</b></div>
          <div><span>量化</span><b>{fabric.quantification_m_per_kg or "缺失"}</b></div>
          <div><span>纸筒空差</span><b>{fabric.tube_plus_allowance_kg_per_roll if fabric.tube_plus_allowance_kg_per_roll is not None else "缺失"}</b></div>
        </div>
        <div class="muted-line">说明：{html.escape(exchange_source)}</div>
        """,
        unsafe_allow_html=True,
    )
    notes = fabric.remarks_en or fabric.remarks_cn
    if notes:
        st.warning(f"数据库备注：{notes}")


def _price_rules(fabric_code: str, price_rules: pd.DataFrame, selected_grade: str | None = None) -> None:
    if price_rules.empty or not fabric_code:
        return
    subset = price_rules[price_rules["fabric_code"].astype(str).str.strip().str.casefold() == str(fabric_code).casefold()]
    if selected_grade and "quality_grade_en" in subset.columns:
        grade = selected_grade.strip().casefold()
        grade_subset = subset[
            subset["quality_grade_en"].fillna("").astype(str).str.casefold().eq(grade)
            | subset.get("quality_grade_cn", pd.Series("", index=subset.index)).fillna("").astype(str).str.casefold().eq(grade)
        ]
        if not grade_subset.empty:
            subset = grade_subset
    if subset.empty:
        return
    with st.expander("价格规则", expanded=False):
        keep = [
            c
            for c in [
                "quality_grade_en",
                "color_codes",
                "price_adjustment_cn",
                "bulk_price_rmb_per_kg",
                "net_price_rmb_per_kg",
                "remarks_en",
            ]
            if c in subset.columns
        ]
        st.dataframe(subset[keep], width="stretch", hide_index=True)


def _ocr_fabric_image(uploaded) -> tuple[str, str]:
    try:
        from PIL import Image
    except ImportError:
        return "", "缺少 pillow，无法做截图 OCR。可以先把截图文字复制到文本框。"
    try:
        uploaded.seek(0)
        image = Image.open(uploaded)
        text = _ocr_image_to_text(image)
    except Exception as exc:  # noqa: BLE001
        return "", f"截图 OCR 失败：{exc}。可以先把截图文字复制到文本框。"
    text = text.strip()
    if not text:
        return "", "截图 OCR 没有识别出文字，请换更清晰截图或手动粘贴。"
    return text, "已识别截图文字，请检查后套用。"


def _ocr_uploaded_images(uploaded_files, default_price: float | None = None) -> tuple[str, str]:
    try:
        from PIL import Image
    except ImportError:
        return "", "缺少 pillow，无法读取拖入的图片。"

    files = list(uploaded_files or [])
    if not files:
        return "", "还没有拖入码单图片。"

    items = []
    failed = []
    for index, uploaded in enumerate(files, start=1):
        name = getattr(uploaded, "name", f"截图{index}")
        try:
            uploaded.seek(0)
            image = Image.open(uploaded)
            text = _ocr_roll_image_to_text(image, default_price=default_price).strip()
        except Exception as exc:  # noqa: BLE001
            failed.append(f"{name}: {exc}")
            continue
        if text:
            items.append((name, text))

    combined = _combine_ocr_texts(items)
    if not combined:
        detail = f"；失败：{' / '.join(failed)}" if failed else ""
        return "", f"拖入图片没有识别出文字，请换更清晰截图或手动粘贴文字{detail}。"
    message = f"已识别 {len(items)} 张码单图片，结果已合并到文本框。"
    if failed:
        message += f" 其中 {len(failed)} 张失败，请人工检查。"
    return combined, message


def _ocr_clipboard_image(prefer_roll_parser: bool = False, default_price: float | None = None) -> tuple[str, str]:
    try:
        from PIL import Image, ImageGrab
    except ImportError:
        return "", "缺少 pillow，无法读取剪贴板截图。"
    try:
        clipboard = ImageGrab.grabclipboard()
    except Exception as exc:  # noqa: BLE001
        return "", f"读取剪贴板失败：{exc}"
    if clipboard is None:
        return "", "剪贴板里没有图片。请先截图复制，再点击读取剪贴板截图。"
    try:
        items = []
        if isinstance(clipboard, Image.Image):
            text = _ocr_roll_image_to_text(clipboard, default_price=default_price) if prefer_roll_parser else _ocr_image_to_text(clipboard)
            items.append(("剪贴板截图", text.strip()))
        elif isinstance(clipboard, list) and clipboard:
            for index, path in enumerate(clipboard, start=1):
                image_path = Path(path)
                if not image_path.exists():
                    continue
                try:
                    image = Image.open(image_path)
                    if prefer_roll_parser:
                        text = _ocr_roll_image_to_text(image, default_price=default_price).strip()
                    else:
                        text = _ocr_image_to_text(image).strip()
                except Exception:
                    continue
                if text:
                    items.append((image_path.name or f"剪贴板图片{index}", text))
        else:
            return "", "剪贴板内容不是图片。请先截图复制，再点击读取剪贴板截图。"
        text = _combine_ocr_texts(items)
    except Exception as exc:  # noqa: BLE001
        return "", f"剪贴板截图 OCR 失败：{exc}"
    if not text:
        return "", "剪贴板截图没有识别出文字，请换更清晰截图或手动粘贴文字。"
    count = len(items)
    return text, f"已读取并识别 {count} 张剪贴板截图，请检查后解析到明细表。"


def _combine_ocr_texts(items: list[tuple[str, str]]) -> str:
    chunks = []
    for index, (name, text) in enumerate(items, start=1):
        clean = str(text or "").strip()
        if not clean:
            continue
        chunks.append(f"--- 码单截图 {index}: {name} ---\n{clean}")
    return "\n\n".join(chunks).strip()


def _ocr_roll_image_to_text(image, default_price: float | None = None) -> str:
    structured = _company_roll_image_to_text(image, default_price=default_price)
    if structured:
        return structured
    return _ocr_image_to_text(image)


def _company_roll_image_to_text(image, default_price: float | None = None) -> str:
    lines = _rapidocr_lines(image)
    if not lines or not _looks_like_company_roll_image(lines):
        return ""

    width, height = image.size
    header_y = _header_y(lines, height)
    summary_y = _summary_y(lines, header_y, height)
    body = [line for line in lines if header_y + 8 <= line["cy"] <= summary_y]
    if not body:
        return ""

    def in_col(line: dict[str, object], left: float, right: float) -> bool:
        return width * left <= float(line["cx"]) < width * right

    weight_lines = [
        line
        for line in body
        if in_col(line, 0.32, 0.705) and len(_roll_weight_numbers(str(line["text"]))) >= 1
    ]
    weight_lines.sort(key=lambda line: float(line["cy"]))

    lot_lines = [
        line
        for line in body
        if in_col(line, 0.19, 0.325) and re.fullmatch(r"[A-Z]{1,5}\d{6,}[\w-]*", str(line["text"]).strip(), flags=re.I)
    ]
    lot_lines.sort(key=lambda line: float(line["cy"]))
    if not lot_lines or not weight_lines:
        return ""

    anchors = []
    for lot in lot_lines:
        y = float(lot["cy"])
        item = _nearby_cell_text(body, 0.00, 0.075, y, width)
        color = _nearby_cell_text(body, 0.075, 0.195, y, width)
        rolls = _nearby_number(body, 0.805, 0.86, y, width, integer=True)
        ocr_price = _nearby_number(body, 0.86, 0.91, y, width, integer=False)
        price = default_price or ocr_price
        anchors.append(
            {
                "item": item,
                "color": color,
                "lot": str(lot["text"]).strip(),
                "rolls": int(rolls) if rolls else 0,
                "price": price,
                "cy": y,
            }
        )

    rows = []
    pointer = 0
    for index, anchor in enumerate(anchors):
        expected = int(anchor["rolls"] or 0)
        next_anchor_y = float(anchors[index + 1]["cy"]) if index + 1 < len(anchors) else summary_y
        pieces = []
        count = 0
        while pointer < len(weight_lines):
            line = weight_lines[pointer]
            y = float(line["cy"])
            if not expected and y > next_anchor_y:
                break
            numbers = _roll_weight_numbers(str(line["text"]))
            if not numbers:
                pointer += 1
                continue
            pieces.append(", ".join(_format_ocr_number(number) for number in numbers))
            count += len(numbers)
            pointer += 1
            if expected and count >= expected:
                break
        if pieces and anchor["color"] and anchor["lot"]:
            rows.append(
                "\t".join(
                    [
                        str(anchor["item"]).strip(),
                        str(anchor["color"]).strip(),
                        str(anchor["lot"]).strip(),
                        ", ".join(pieces),
                        _format_ocr_number(anchor["price"]) if anchor["price"] else "",
                    ]
                )
            )

    return "\n".join(rows).strip()


def _rapidocr_lines(image) -> list[dict[str, object]]:
    try:
        import numpy as np
        from rapidocr_onnxruntime import RapidOCR
    except ImportError:
        return []
    result, _ = _rapidocr_engine()(np.array(image.convert("RGB")))
    lines = []
    for item in result or []:
        box, text, score = item
        xs = [point[0] for point in box]
        ys = [point[1] for point in box]
        lines.append(
            {
                "text": str(text).strip(),
                "score": float(score),
                "cx": sum(xs) / len(xs),
                "cy": sum(ys) / len(ys),
                "x1": min(xs),
                "x2": max(xs),
                "y1": min(ys),
                "y2": max(ys),
            }
        )
    return [line for line in lines if line["text"]]


@st.cache_resource(show_spinner=False)
def _rapidocr_engine():
    from rapidocr_onnxruntime import RapidOCR

    return RapidOCR()


def _looks_like_company_roll_image(lines: list[dict[str, object]]) -> bool:
    text = "\n".join(str(line["text"]) for line in lines)
    required = ["款号", "颜色", "缸号", "细码数量", "条数"]
    return sum(label in text for label in required) >= 4 and ("Yoooni" in text or "客户名称" in text or "销售单号" in text)


def _header_y(lines: list[dict[str, object]], height: int) -> float:
    candidates = [float(line["cy"]) for line in lines if str(line["text"]) in {"款号", "颜色", "缸号", "细码数量", "数量", "条数"}]
    return sum(candidates) / len(candidates) if candidates else height * 0.27


def _summary_y(lines: list[dict[str, object]], header_y: float, height: int) -> float:
    candidates = [float(line["cy"]) for line in lines if float(line["cy"]) > header_y and "细码合计" in str(line["text"])]
    return min(candidates) - 8 if candidates else height * 0.72


def _nearby_cell_text(lines: list[dict[str, object]], left: float, right: float, y: float, width: int, tolerance: float = 22) -> str:
    candidates = [
        line
        for line in lines
        if width * left <= float(line["cx"]) < width * right and abs(float(line["cy"]) - y) <= tolerance
    ]
    if not candidates:
        return ""
    candidates.sort(key=lambda line: (abs(float(line["cy"]) - y), -float(line["score"])))
    return str(candidates[0]["text"]).strip()


def _nearby_number(
    lines: list[dict[str, object]],
    left: float,
    right: float,
    y: float,
    width: int,
    integer: bool,
    tolerance: float = 22,
) -> float | None:
    text = _nearby_cell_text(lines, left, right, y, width, tolerance=tolerance)
    numbers = re.findall(r"\d+(?:\.\d+)?", text.replace(",", ""))
    if not numbers:
        return None
    value = float(numbers[0])
    return int(value) if integer else value


def _roll_weight_numbers(text: str) -> list[float]:
    values = []
    for match in re.finditer(r"\d+(?:\.\d+)?", str(text).replace("，", ",")):
        value = float(match.group(0))
        if 3 <= value <= 80:
            values.append(value)
    return values


def _format_ocr_number(value: object) -> str:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return ""
    if number.is_integer():
        return str(int(number))
    return f"{number:.2f}".rstrip("0").rstrip(".")


def _ocr_image_to_text(image) -> str:
    rapid_lines = _rapidocr_lines(image)
    if rapid_lines:
        rapid_lines.sort(key=lambda line: (round(float(line["cy"]) / 12), float(line["cx"])))
        return "\n".join(str(line["text"]) for line in rapid_lines).strip()
    try:
        import pytesseract
        from PIL import ImageOps
    except ImportError as exc:
        raise RuntimeError("缺少 OCR 引擎，无法识别截图。请安装 rapidocr_onnxruntime，或先把截图文字复制到文本框。") from exc
    source = image.convert("L")
    width, height = source.size
    scale = 2 if max(width, height) < 2200 else 1
    prepared = ImageOps.autocontrast(source).resize((max(width * scale, width), max(height * scale, height)))
    threshold = prepared.point(lambda pixel: 255 if pixel > 180 else 0)
    variants = [prepared, threshold]
    configs = ["--psm 6", "--psm 11", "--psm 12"]
    results = []
    for variant in variants:
        for config in configs:
            try:
                text = pytesseract.image_to_string(variant, lang="eng+chi_sim", config=config).strip()
            except Exception:
                text = pytesseract.image_to_string(variant, lang="eng", config=config).strip()
            if text:
                results.append(text)
    return max(results, key=_ocr_score, default="").strip()


def _ocr_score(text: str) -> int:
    value = str(text or "")
    number_count = len(re.findall(r"\d+(?:\.\d+)?", value))
    line_count = len([line for line in value.splitlines() if line.strip()])
    return len(value) + number_count * 8 + line_count * 3


def _parse_fabric_text(text: str, fallback: FabricRecord | None) -> FabricRecord:
    base = fallback or FabricRecord(fabric_code="")
    value = str(text or "")
    code = (
        _extract_between(
            value,
            ["面料编号", "fabric code", "quality code", "ITEM", "code"],
            ["comp", "composition", "成分", "weight", "克重", "width", "幅宽", "finish", "country"],
        )
        or base.fabric_code
    )
    grade = _extract_grade(value) or base.quality_grade_en
    name_en = _extract_value(value, [r"(?:英文品名|fabric\s*name\s*en|fabric\s*name|品名)\s*[:：]?\s*([^\n；;]+)"]) or base.fabric_name_en
    name_cn = _extract_value(value, [r"(?:中文品名|fabric\s*name\s*cn)\s*[:：]?\s*([^\n；;]+)"]) or base.fabric_name_cn
    composition_en = (
        _extract_between(value, ["英文成分", "composition en", "composition", "comp", "成分", "fabrication"], ["weight", "克重", "width", "幅宽", "finish", "country"])
        or base.composition_en
    )
    weight_en = (
        _extract_between(value, ["克重", "weight"], ["width", "幅宽", "finish", "country"])
        or _extract_value(value, [r"(\d+(?:\.\d+)?\s*(?:gsm|g/m2|GSM|G/M2|G))"])
        or base.weight_en
    )
    width_en = (
        _extract_between(value, ["幅宽", "width"], ["weight", "克重", "finish", "country"])
        or _extract_value(value, [r"(\d+(?:\.\d+)?\s*(?:cm|CM|inch|\"))"])
        or base.width_en
    )
    quantification = _extract_float(value, [r"(?:量化|quantification|meter[_ ]?per[_ ]?kg|m/kg)\D*(\d+(?:\.\d+)?)"])
    tube = _extract_float(value, [r"(?:纸筒空差|tube[_ ]?plus|tube.*allowance|空差)\D*(\d+(?:\.\d+)?)"])
    remarks_en = _extract_value(value, [r"(?:备注|remarks?)\s*[:：]?\s*([^\n]+)"]) or base.remarks_en
    return FabricRecord(
        fabric_code=code,
        quality_grade_cn=base.quality_grade_cn,
        quality_grade_en=grade,
        fabric_name_cn=name_cn,
        fabric_name_en=name_en,
        composition_cn=base.composition_cn,
        composition_en=composition_en,
        width_cn=base.width_cn,
        width_en=width_en,
        weight_cn=base.weight_cn,
        weight_en=weight_en,
        quantification_m_per_kg=quantification if quantification is not None else base.quantification_m_per_kg,
        tube_plus_allowance_kg_per_roll=tube if tube is not None else base.tube_plus_allowance_kg_per_roll,
        reference_roll_weight_kg=base.reference_roll_weight_kg,
        remarks_cn=base.remarks_cn,
        remarks_en=remarks_en,
        raw=base.raw,
    )


def _fabric_candidate_editor(fabric: FabricRecord, source_text: str) -> FabricRecord:
    st.caption("识别结果可在这里修改。下面按钮会使用你修改后的这一行。")
    row = {
        "fabric_code": fabric.fabric_code,
        "quality_grade_en": fabric.quality_grade_en,
        "fabric_name_en": fabric.fabric_name_en,
        "composition_en": fabric.composition_en,
        "weight_en": fabric.weight_en,
        "width_en": fabric.width_en,
        "quantification_m_per_kg": fabric.quantification_m_per_kg,
        "tube_plus_allowance_kg_per_roll": fabric.tube_plus_allowance_kg_per_roll,
        "remarks_en": fabric.remarks_en,
    }
    editor_key = f"fabric_candidate_{abs(hash(source_text))}_{len(source_text)}"
    edited = st.data_editor(
        pd.DataFrame([row]),
        num_rows="fixed",
        width="stretch",
        height=120,
        key=editor_key,
        column_config={
            "fabric_code": st.column_config.TextColumn("面料编号"),
            "quality_grade_en": st.column_config.TextColumn("质量等级"),
            "fabric_name_en": st.column_config.TextColumn("英文品名"),
            "composition_en": st.column_config.TextColumn("英文成分"),
            "weight_en": st.column_config.TextColumn("克重"),
            "width_en": st.column_config.TextColumn("幅宽"),
            "quantification_m_per_kg": st.column_config.NumberColumn("量化", min_value=0.0, step=0.01, format="%.4f"),
            "tube_plus_allowance_kg_per_roll": st.column_config.NumberColumn("纸筒空差", min_value=0.0, step=0.1, format="%.2f"),
            "remarks_en": st.column_config.TextColumn("备注"),
        },
        hide_index=True,
    )
    df = pd.DataFrame(edited)
    if df.empty:
        return fabric
    item = df.iloc[0].to_dict()
    return FabricRecord(
        fabric_code=str(item.get("fabric_code") or "").strip(),
        quality_grade_cn=fabric.quality_grade_cn,
        quality_grade_en=str(item.get("quality_grade_en") or "").strip(),
        fabric_name_cn=fabric.fabric_name_cn,
        fabric_name_en=str(item.get("fabric_name_en") or "").strip(),
        composition_cn=fabric.composition_cn,
        composition_en=str(item.get("composition_en") or "").strip(),
        width_cn=fabric.width_cn,
        width_en=str(item.get("width_en") or "").strip(),
        weight_cn=fabric.weight_cn,
        weight_en=str(item.get("weight_en") or "").strip(),
        quantification_m_per_kg=_to_optional_float(item.get("quantification_m_per_kg")),
        tube_plus_allowance_kg_per_roll=_to_optional_float(item.get("tube_plus_allowance_kg_per_roll")),
        reference_roll_weight_kg=fabric.reference_roll_weight_kg,
        remarks_cn=fabric.remarks_cn,
        remarks_en=str(item.get("remarks_en") or "").strip(),
        raw=fabric.raw,
    )


def _apply_fabric_to_editor(fabric: FabricRecord, key_prefix: str) -> None:
    st.session_state[f"{key_prefix}_code"] = fabric.fabric_code
    st.session_state[f"{key_prefix}_name_en"] = fabric.fabric_name_en
    st.session_state[f"{key_prefix}_name_cn"] = fabric.fabric_name_cn
    st.session_state[f"{key_prefix}_grade_en"] = fabric.quality_grade_en
    st.session_state[f"{key_prefix}_comp_en"] = fabric.composition_en
    st.session_state[f"{key_prefix}_weight_en"] = fabric.weight_en
    st.session_state[f"{key_prefix}_width_en"] = fabric.width_en
    st.session_state[f"{key_prefix}_quantification"] = float(fabric.quantification_m_per_kg or 0.0)
    st.session_state[f"{key_prefix}_tube"] = float(fabric.tube_plus_allowance_kg_per_roll or 0.0)
    st.session_state[f"{key_prefix}_remarks_en"] = fabric.remarks_en


def _validate_fabric_record(fabric: FabricRecord) -> list[str]:
    errors: list[str] = []
    if not fabric.fabric_code.strip():
        errors.append("面料编号不能为空")
    if not fabric.composition_en.strip():
        errors.append("英文成分不能为空")
    if not fabric.weight_en.strip():
        errors.append("克重不能为空")
    if not fabric.width_en.strip():
        errors.append("幅宽不能为空")
    if not fabric.quantification_m_per_kg or fabric.quantification_m_per_kg <= 0:
        errors.append("量化 m/kg 必须大于 0")
    if fabric.tube_plus_allowance_kg_per_roll is None:
        errors.append("纸筒空差不能为空")
    return errors


def _price_hint(fabric_code: str, grade: str, price_rules: pd.DataFrame) -> str:
    if price_rules.empty or not fabric_code:
        return ""
    subset = price_rules[price_rules["fabric_code"].astype(str).str.strip().str.casefold() == str(fabric_code).casefold()]
    if subset.empty:
        return ""
    if grade and "quality_grade_en" in subset.columns:
        matched = subset[
            subset["quality_grade_en"].fillna("").astype(str).str.casefold().eq(grade.strip().casefold())
            | subset.get("quality_grade_cn", pd.Series("", index=subset.index)).fillna("").astype(str).str.casefold().eq(grade.strip().casefold())
        ]
        if not matched.empty:
            subset = matched
    prices = pd.to_numeric(subset.get("net_price_rmb_per_kg"), errors="coerce").dropna()
    if prices.empty:
        return ""
    min_price = float(prices.min())
    max_price = float(prices.max())
    if min_price == max_price:
        return f"当前质量等级关联净价：RMB {min_price:.2f}/KG（仅作价格规则参考）"
    return f"当前质量等级关联净价：RMB {min_price:.2f}-{max_price:.2f}/KG（按色号规则不同）"


def _best_fabric_default(fabric_codes: list[str]) -> str:
    for code in ["6373D", "E2500292"]:
        if code in fabric_codes:
            return code
    return fabric_codes[0] if fabric_codes else ""


def _first_number(*values: object) -> float | None:
    for value in values:
        try:
            if pd.isna(value):
                continue
            number = float(value)
        except (TypeError, ValueError):
            continue
        if number > 0:
            return number
    return None


def _extract_value(text: str, patterns: list[str]) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I | re.S)
        if match:
            return _clean_extracted(match.group(1))
    return ""


def _extract_between(text: str, labels: list[str], stops: list[str]) -> str:
    label_pattern = "|".join(re.escape(label) for label in labels)
    stop_pattern = "|".join(re.escape(stop) for stop in stops)
    pattern = rf"(?:{label_pattern})\s*[:：]?\s*(.*?)(?=(?:{stop_pattern})\s*[:：]?|$)"
    match = re.search(pattern, text, flags=re.I | re.S)
    return _clean_extracted(match.group(1)) if match else ""


def _extract_float(text: str, patterns: list[str]) -> float | None:
    for pattern in patterns:
        value = _extract_value(text, [pattern])
        if value:
            try:
                return float(value)
            except ValueError:
                continue
    return None


def _to_optional_float(value: object) -> float | None:
    try:
        if pd.isna(value):
            return None
        text = str(value).strip()
        if not text:
            return None
        return float(text)
    except (TypeError, ValueError):
        return None


def _clean_extracted(value: str) -> str:
    text = re.sub(r"\s+", " ", str(value or "").strip())
    text = re.sub(r"^(?:[:：;-]+)\s*", "", text)
    text = re.sub(r"\s*(?:[:：;-]+)$", "", text)
    return text.strip()


def _extract_grade(text: str) -> str:
    lower = text.casefold()
    if "first grade" in lower or "一等品" in text:
        return "First Grade"
    if "qualified grade" in lower or "合格品" in text:
        return "Qualified Grade"
    return _extract_value(text, [r"(?:质量等级|quality\s*grade)\s*[:：]?\s*([^\n；;]+)"])


def _fabric_widget_prefix(fabric_code: str, grade: str) -> str:
    raw = f"{fabric_code}_{grade}"
    safe = re.sub(r"[^0-9A-Za-z_]+", "_", raw).strip("_")
    return f"fabric_edit_{safe or 'manual'}"


def _read_clipboard_pdf() -> tuple[Path | None, str]:
    paths = _clipboard_file_paths()
    pdfs = [path for path in paths if path.suffix.lower() == ".pdf" and path.exists()]
    if pdfs:
        return pdfs[0], ""
    try:
        import tkinter as tk

        root = tk.Tk()
        root.withdraw()
        text = root.clipboard_get()
        root.destroy()
        candidate = Path(str(text).strip().strip('"'))
        if candidate.exists() and candidate.suffix.lower() == ".pdf":
            return candidate, ""
    except Exception:
        pass
    return None, "剪贴板里没有 PDF 文件。请在资源管理器复制 PDF，或把 PDF 拖进上传框。"


def _clipboard_file_paths() -> list[Path]:
    if not hasattr(ctypes, "windll"):
        return []
    user32 = ctypes.windll.user32
    shell32 = ctypes.windll.shell32
    CF_HDROP = 15
    paths: list[Path] = []

    user32.OpenClipboard.argtypes = [wintypes.HWND]
    user32.OpenClipboard.restype = wintypes.BOOL
    user32.CloseClipboard.argtypes = []
    user32.CloseClipboard.restype = wintypes.BOOL
    user32.IsClipboardFormatAvailable.argtypes = [wintypes.UINT]
    user32.IsClipboardFormatAvailable.restype = wintypes.BOOL
    user32.GetClipboardData.argtypes = [wintypes.UINT]
    user32.GetClipboardData.restype = wintypes.HANDLE
    shell32.DragQueryFileW.argtypes = [wintypes.HANDLE, wintypes.UINT, wintypes.LPWSTR, wintypes.UINT]
    shell32.DragQueryFileW.restype = wintypes.UINT

    opened = False
    try:
        if not user32.OpenClipboard(None):
            return paths
        opened = True
        if not user32.IsClipboardFormatAvailable(CF_HDROP):
            return paths
        handle = user32.GetClipboardData(CF_HDROP)
        if not handle:
            return paths
        count = shell32.DragQueryFileW(handle, 0xFFFFFFFF, None, 0)
        for index in range(count):
            length = shell32.DragQueryFileW(handle, index, None, 0)
            if length <= 0:
                continue
            buffer = ctypes.create_unicode_buffer(length + 1)
            shell32.DragQueryFileW(handle, index, buffer, length + 1)
            if buffer.value:
                paths.append(Path(buffer.value))
    except (OSError, ValueError):
        return []
    finally:
        if opened:
            user32.CloseClipboard()
    return paths


def _file_select(label: str, suffixes: list[str], default_name: str) -> Path | None:
    files = sorted([p for p in BASE_DIR.iterdir() if p.is_file() and any(str(p).lower().endswith(s.lower()) for s in suffixes)])
    options = [""] + [p.name for p in files]
    default_path = BASE_DIR / default_name
    default_index = options.index(default_path.name) if default_path.exists() and default_path.name in options else 0
    selected = st.selectbox(label, options, index=default_index)
    return BASE_DIR / selected if selected else None


def _uploaded_or_current(label: str, types: list[str], current: Path | None) -> Path | None:
    uploaded = st.file_uploader(label, type=types, key=f"upload_{label}")
    if uploaded is None:
        return current
    upload_dir = BASE_DIR / ".streamlit_uploads"
    upload_dir.mkdir(exist_ok=True)
    target = upload_dir / uploaded.name
    target.write_bytes(uploaded.getbuffer())
    return target


def _default_buyer(template: Path | None, module: str) -> str:
    if not template or not template.exists():
        return ""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(template, data_only=True)
        ws = wb.active
        value = ws["A11"].value if module == "invoice" else ws["G5"].value
        text = str(value or "").strip()
        if text.upper().startswith("BUYER:"):
            text = text.split(":", 1)[1].strip()
        return text
    except Exception:
        return ""


def _grade_options(df: pd.DataFrame, fabric_code: str) -> list[str]:
    subset = df[df["fabric_code"].astype(str).str.strip().str.casefold() == str(fabric_code).casefold()]
    options: list[str] = []
    for _, row in subset.iterrows():
        grade = str(row.get("quality_grade_en") or row.get("quality_grade_cn") or "").strip()
        if grade and grade not in options:
            options.append(grade)
    return options or [""]


def _style() -> None:
    st.markdown(
        """
        <style>
        :root {
            --bg: #f6f7f8;
            --surface: #ffffff;
            --text: #17201c;
            --muted: #66716c;
            --line: #dce2df;
            --accent: #177a62;
            --accent-soft: #e5f3ee;
            --danger: #b42318;
            --danger-soft: #fee4e2;
        }
        .stApp { background: var(--bg); color: var(--text); }
        .block-container { max-width: 1180px; padding-top: 1.2rem; padding-bottom: 3rem; }
        h1, h2, h3 { letter-spacing: 0; }
        label { color: var(--muted) !important; font-size: .78rem !important; font-weight: 650 !important; }
        .hero { padding: 26px 0 18px; }
        .hero h1, .module-title h1 { margin: 0; font-size: 2.2rem; line-height: 1.15; font-weight: 780; }
        .hero p, .module-title p { margin: 8px 0 0; color: var(--muted); font-size: .98rem; }
        .entry-card {
            min-height: 180px; border: 1px solid var(--line); border-radius: 14px;
            background: var(--surface); padding: 24px; margin-bottom: 12px;
        }
        .entry-card h2 { margin: 0 0 12px; font-size: 1.65rem; }
        .entry-card p { color: var(--muted); max-width: 420px; margin: 0 0 34px; }
        .entry-card span { color: var(--accent); font-weight: 750; font-size: .84rem; }
        .section-label {
            margin: 26px 0 12px; font-weight: 780; font-size: 1.05rem; color: var(--text);
        }
        .summary-row, .fabric-line, .tiny-status, .import-box, .doc-box {
            display: grid; gap: 1px; border: 1px solid var(--line); border-radius: 12px;
            overflow: hidden; background: var(--line); margin: 16px 0;
        }
        .summary-row { grid-template-columns: 1fr 1.45fr .75fr .55fr .75fr .62fr; }
        .fabric-line { grid-template-columns: 1.2fr .8fr .8fr .55fr .65fr; }
        .tiny-status { grid-template-columns: 1.5fr .5fr .5fr; margin-top: 28px; }
        .import-box { grid-template-columns: 1fr 1fr 1.6fr .6fr .9fr; }
        .doc-box { grid-template-columns: 1fr 1fr 1.2fr; }
        .summary-row > div, .fabric-line > div, .tiny-status > div, .import-box > div, .doc-box > div {
            background: var(--surface); padding: 13px 14px; min-width: 0;
        }
        .summary-row span, .fabric-line span, .tiny-status span, .import-box span, .doc-box span {
            display: block; color: var(--muted); font-size: .72rem; font-weight: 700; margin-bottom: 5px;
        }
        .summary-row b, .fabric-line b, .tiny-status b, .import-box b, .doc-box b {
            display: block; color: var(--text); font-size: .92rem; font-weight: 760;
            white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
        }
        .summary-row b.good { color: var(--accent); }
        .summary-row b.bad { color: var(--danger); }
        .muted-line { color: var(--muted); font-size: .78rem; margin-top: -8px; }
        .sub-label { margin: 0 0 8px; font-weight: 760; color: var(--text); }
        .stButton > button, .stDownloadButton > button {
            border-radius: 9px; border: 1px solid var(--line); font-weight: 700; min-height: 2.55rem;
        }
        .stButton > button[kind="primary"] { background: var(--accent); border-color: var(--accent); }
        div[data-baseweb="input"], div[data-baseweb="select"] > div { border-radius: 9px !important; }
        @media (max-width: 900px) {
            .summary-row, .fabric-line, .tiny-status, .import-box, .doc-box { grid-template-columns: 1fr; }
            .hero h1, .module-title h1 { font-size: 1.8rem; }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    main()
