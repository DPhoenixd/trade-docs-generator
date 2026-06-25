from __future__ import annotations

import json
import re
import shutil
import uuid
from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import load_workbook

from .calculations import (
    INVOICE_COLUMNS,
    ROLL_COLUMNS,
    apply_gross_weight,
    calculate_invoice_lines,
    calculate_rolls,
)
from .excel_writer import write_packing_list, write_proforma_invoice
from .fabric_db import fabric_record_from_row, find_fabric, load_fabric_master
from .models import FabricRecord, OrderInfo
from .numbering import suggest_doc_number
from .paths import app_data_dir, default_fabric_database, project_root


BASE_DIR = project_root()
DATA_DIR = app_data_dir()
RUNTIME_DIR = DATA_DIR / ".pipl_runtime"
SESSION_DIR = RUNTIME_DIR / "sessions"
OUTPUT_DIR = DATA_DIR / "outputs"

DEFAULT_PI_TEMPLATE = Path(
    "D:/DyrusWok/外贸/COTTEX&TESSELATION/#6529/POUT26VE0002814A -6529/6.08-P.I- P.L/6.08 P.I-POUT26VE0002814A -6529.xlsx"
)
DEFAULT_PL_TEMPLATE = Path(
    "D:/DyrusWok/外贸/COTTEX&TESSELATION/#6529/POUT26VE0002814A -6529/6.08-P.I- P.L/6.08Packing List-POUT26VE0002814A-6529-20260608.xlsx"
)
DEFAULT_FABRIC_DB = default_fabric_database()


def analyze_files(saved_files: list[Path] | None = None) -> dict[str, Any]:
    SESSION_DIR.mkdir(parents=True, exist_ok=True)
    session_id = uuid.uuid4().hex[:12]
    session_path = SESSION_DIR / session_id
    session_path.mkdir(parents=True, exist_ok=True)
    if not saved_files:
        manifest = {
            "session_id": session_id,
            "paths": {},
            "order_images": [],
            "order_image_rows": [],
            "invoice_input": [],
            "roll_input": [],
        }
        _write_manifest(session_path, manifest)
        return _empty_analysis(session_id)

    paths = _default_paths()
    order_images: list[Path] = []
    for source in saved_files or []:
        target = session_path / _safe_filename(source.name)
        shutil.copy2(source, target)
        kind = _classify_file(target)
        if kind == "order_image":
            order_images.append(target)
        elif kind:
            paths[kind] = target

    template_invoice_input = _invoice_rows_from_template(paths.get("pi_template"))
    roll_input = _roll_rows_from_packing_template(paths.get("pl_template"))
    template_defaults = _invoice_template_defaults(paths.get("pi_template"))
    pl_defaults = _packing_template_defaults(paths.get("pl_template"))

    fabric_code = (
        template_defaults.get("fabric_code")
        or pl_defaults.get("fabric_code")
        or _filename_fabric_code(paths.get("pi_template"))
        or _filename_fabric_code(paths.get("pl_template"))
        or ""
    )
    fabric_df = _load_fabric_df(paths.get("fabric_db"))
    fabric = None
    if fabric_code and "fabric_code" in fabric_df.columns:
        fabric = find_fabric(fabric_df, fabric_code, "Qualified Grade") or find_fabric(fabric_df, fabric_code)
    if fabric is None and fabric_code and (template_defaults or pl_defaults):
        fabric = _fabric_from_template(fabric_code, template_defaults, pl_defaults)

    order_image_rows = _parse_order_images(order_images)
    order_image_errors = [row for row in order_image_rows if row.get("parse_error")]
    order_image_rows = [row for row in order_image_rows if not row.get("parse_error")]
    invoice_input = _invoice_rows_from_order_images(order_image_rows, fabric, template_invoice_input) if order_image_rows else template_invoice_input
    invoice_lines = _compute_invoice(fabric, invoice_input)
    computed_rolls, packing_summary = _compute_packing(fabric, roll_input)
    order = _build_order(template_defaults, pl_defaults, fabric_code, invoice_lines)
    issues = _build_issues(order, invoice_lines, computed_rolls, packing_summary, fabric, paths)

    manifest = {
        "session_id": session_id,
        "paths": {key: str(value) for key, value in paths.items() if value},
        "order_images": [str(path) for path in order_images],
        "order_image_rows": order_image_rows,
        "order_image_errors": order_image_errors,
        "invoice_input": _records(invoice_input),
        "roll_input": _records(roll_input),
        "order": asdict(order),
        "fabric": asdict(fabric) if fabric else None,
    }
    _write_manifest(session_path, manifest)

    return {
        "session_id": session_id,
        "files": _file_status(paths, order_images),
        "order_image_rows": order_image_rows,
        "order_image_errors": order_image_errors,
        "invoice_input": _records(invoice_input),
        "roll_input": _records(roll_input),
        "order": asdict(order),
        "fabric": _fabric_payload(fabric),
        "fabric_record": asdict(fabric) if fabric else None,
        "invoice_rows": _records(invoice_lines),
        "roll_rows": _records(computed_rolls),
        "packing_summary": _records(packing_summary),
        "issues": issues,
        "can_generate": _can_generate(invoice_lines, packing_summary, fabric, paths),
        "totals": _totals(invoice_lines, computed_rolls, packing_summary),
    }


def generate_documents(payload: dict[str, Any]) -> dict[str, Any]:
    session_id = str(payload.get("session_id") or "")
    session_path = SESSION_DIR / session_id
    manifest = _read_manifest(session_path)
    paths = {key: Path(value) for key, value in manifest.get("paths", {}).items()}
    if not paths.get("pi_template") or not paths.get("pl_template"):
        raise ValueError("缺少 P.I 或 Packing List 模板，请先上传模板。")

    order_data = {**manifest.get("order", {}), **(payload.get("order") or {})}
    order_data["order_date"] = _today().strftime("%Y-%m-%d")
    order_data["pi_no"] = suggest_doc_number(today=_today())
    order_data["ci_no"] = order_data["pi_no"]
    order_data["output_dir"] = OUTPUT_DIR

    fabric_data = payload.get("fabric") or manifest.get("fabric") or {}
    if not fabric_data:
        raise ValueError("缺少面料数据，请先选择或上传面料数据库。")
    fabric = FabricRecord(**fabric_data)
    invoice_input = pd.DataFrame(payload.get("invoice_input") or manifest.get("invoice_input") or [], columns=INVOICE_COLUMNS)
    roll_input = pd.DataFrame(payload.get("roll_input") or manifest.get("roll_input") or [], columns=ROLL_COLUMNS)
    invoice_lines = _compute_invoice(fabric, invoice_input)
    computed_rolls, packing_summary = _compute_packing(fabric, roll_input)
    total_amount = float(invoice_lines["amount_usd"].sum()) if not invoice_lines.empty else 0.0
    order_data["advance_payment_usd"] = round(total_amount * 0.3, 2)
    order = OrderInfo(**_order_kwargs(order_data))

    if not _can_generate(invoice_lines, packing_summary, fabric, paths):
        raise ValueError("当前数据不足，无法生成。请检查模板、面料数据和明细。")

    pi_result = write_proforma_invoice(paths["pi_template"], order, fabric, invoice_lines, quantity_unit="KG")
    pl_result = write_packing_list(paths["pl_template"], order, fabric, computed_rolls, packing_summary)
    manifest["generated"] = {
        "pi": str(pi_result.path),
        "pl": str(pl_result.path),
        "pi_log": str(pi_result.log_path),
        "pl_log": str(pl_result.log_path),
        "generated_at": datetime.now(tz=ZoneInfo("Asia/Shanghai")).isoformat(timespec="seconds"),
    }
    manifest["order"] = asdict(order)
    _write_manifest(session_path, manifest)
    return {
        "session_id": session_id,
        "order": asdict(order),
        "generated": {
            "pi": _download_payload(session_id, "pi", pi_result.path),
            "pl": _download_payload(session_id, "pl", pl_result.path),
        },
        "totals": _totals(invoice_lines, computed_rolls, packing_summary),
    }


def generate_invoice_document(payload: dict[str, Any]) -> dict[str, Any]:
    session_id = str(payload.get("session_id") or "")
    session_path = SESSION_DIR / session_id
    manifest = _read_manifest(session_path)
    paths = {key: Path(value) for key, value in manifest.get("paths", {}).items()}
    if not paths.get("pi_template"):
        raise ValueError("缺少 P.I 模板，请先上传 P.I 模板。")

    order_data = {**manifest.get("order", {}), **(payload.get("order") or {})}
    order_data["order_date"] = _today().strftime("%Y-%m-%d")
    order_data["pi_no"] = suggest_doc_number(today=_today())
    order_data["ci_no"] = order_data["pi_no"]
    order_data["output_dir"] = OUTPUT_DIR

    fabric_data = payload.get("fabric") or manifest.get("fabric") or {}
    if not fabric_data:
        raise ValueError("缺少面料数据，请先选择或上传面料数据库。")
    fabric = FabricRecord(**fabric_data)
    invoice_input = pd.DataFrame(payload.get("invoice_input") or manifest.get("invoice_input") or [], columns=INVOICE_COLUMNS)
    invoice_lines = _compute_invoice(fabric, invoice_input)
    if invoice_lines.empty:
        raise ValueError("P.I 明细为空，请先上传客户截图或手动新增颜色行。")

    total_amount = float(invoice_lines["amount_usd"].sum()) if not invoice_lines.empty else 0.0
    order_data["advance_payment_usd"] = round(total_amount * 0.3, 2)
    order = OrderInfo(**_order_kwargs(order_data))
    pi_result = write_proforma_invoice(paths["pi_template"], order, fabric, invoice_lines, quantity_unit="KG")

    generated = manifest.get("generated") or {}
    generated.update(
        {
            "pi": str(pi_result.path),
            "pi_log": str(pi_result.log_path),
            "generated_at": datetime.now(tz=ZoneInfo("Asia/Shanghai")).isoformat(timespec="seconds"),
        }
    )
    manifest["generated"] = generated
    manifest["order"] = asdict(order)
    _write_manifest(session_path, manifest)
    generated_payload = {"pi": _download_payload(session_id, "pi", pi_result.path)}
    if generated.get("pl"):
        generated_payload["pl"] = _download_payload(session_id, "pl", Path(generated["pl"]))
    return {
        "session_id": session_id,
        "order": asdict(order),
        "generated": generated_payload,
        "totals": _totals(invoice_lines, pd.DataFrame(), pd.DataFrame()),
    }


def download_path(session_id: str, kind: str) -> Path:
    manifest = _read_manifest(SESSION_DIR / session_id)
    generated = manifest.get("generated") or {}
    key = "pi" if kind == "pi" else "pl"
    path = Path(generated.get(key) or "")
    if not path.exists():
        raise FileNotFoundError("文件不存在或尚未生成。")
    return path


def search_fabrics(query: str = "", limit: int = 20) -> dict[str, Any]:
    fabric_df = _load_fabric_df(DEFAULT_FABRIC_DB)
    if fabric_df.empty:
        return {"database": str(DEFAULT_FABRIC_DB), "results": [], "message": "默认面料数据库未找到或为空。"}

    text = str(query or "").strip().casefold()
    candidates = fabric_df.copy()
    if text:
        searchable_columns = [
            column
            for column in [
                "fabric_code",
                "fabric_name_cn",
                "fabric_name_en",
                "quality_grade_cn",
                "quality_grade_en",
                "composition_cn",
                "composition_en",
            ]
            if column in candidates.columns
        ]
        mask = pd.Series(False, index=candidates.index)
        for column in searchable_columns:
            mask = mask | candidates[column].fillna("").astype(str).str.casefold().str.contains(re.escape(text), regex=True)
        candidates = candidates[mask]

    results = []
    for _, row in candidates.head(max(1, min(int(limit or 20), 100))).iterrows():
        fabric = fabric_record_from_row(row)
        color_rule = str(row.get("color_rule_cn", "") or "").strip()
        price_adjustment = str(row.get("price_adjustment_cn", "") or "").strip()
        pricing_note = " / ".join(part for part in [color_rule, price_adjustment] if part and part != "-")
        results.append(
            {
                "label": " - ".join(part for part in [fabric.display_name, pricing_note] if part),
                "pricing_note": pricing_note,
                "summary": _fabric_payload(fabric),
                "record": asdict(fabric),
            }
        )
    return {"database": str(DEFAULT_FABRIC_DB), "results": results}


def _default_paths() -> dict[str, Path]:
    paths: dict[str, Path] = {}
    if DEFAULT_FABRIC_DB.exists():
        paths["fabric_db"] = DEFAULT_FABRIC_DB
    return paths


def _empty_analysis(session_id: str) -> dict[str, Any]:
    return {
        "session_id": session_id,
        "files": _file_status({}, []),
        "order_image_rows": [],
        "order_image_errors": [],
        "invoice_input": [],
        "roll_input": [],
        "order": {},
        "fabric": None,
        "invoice_rows": [],
        "roll_rows": [],
        "packing_summary": [],
        "issues": [],
        "can_generate": False,
        "totals": _totals(pd.DataFrame(), pd.DataFrame(), pd.DataFrame()),
    }


def _classify_file(path: Path) -> str | None:
    name = path.name.casefold()
    if path.suffix.lower() in {".png", ".jpg", ".jpeg", ".webp", ".bmp"}:
        return "order_image"
    if path.suffix.lower() in {".xlsx", ".xlsm", ".csv"} and ("fabric" in name or "price" in name or "0428" in name):
        return "fabric_db"
    if path.suffix.lower() in {".xlsx", ".xlsm"} and ("packing" in name or "p.l" in name or "pl-" in name):
        return "pl_template"
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return "pi_template"
    return None


def _parse_order_images(paths: list[Path]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for path in paths:
        try:
            rows.extend(_parse_order_image(path))
        except Exception as exc:  # noqa: BLE001
            rows.append({"source_file": path.name, "parse_error": str(exc)})
    return rows


def _parse_order_image(path: Path) -> list[dict[str, Any]]:
    try:
        import cv2
        import numpy as np
        from rapidocr_onnxruntime import RapidOCR
    except ImportError as exc:
        raise RuntimeError("缺少截图 OCR 依赖：opencv-python 或 rapidocr_onnxruntime") from exc

    image = _read_cv_image(path, cv2, np)
    if image is None:
        return []
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    _, threshold = cv2.threshold(gray, 200, 255, cv2.THRESH_BINARY_INV)
    height, width = threshold.shape
    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (max(20, width // 25), 1))
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(10, height // 12)))
    horizontal = cv2.morphologyEx(threshold, cv2.MORPH_OPEN, horizontal_kernel)
    vertical = cv2.morphologyEx(threshold, cv2.MORPH_OPEN, vertical_kernel)
    ys = _merge_positions(np.where(horizontal.sum(axis=1) > width * 255 * 0.25)[0])
    xs = _merge_positions(np.where(vertical.sum(axis=0) > height * 255 * 0.25)[0])
    if len(xs) < 2 or len(ys) < 2:
        return _parse_order_image_from_ocr_lines(path)

    engine = RapidOCR()
    result, _ = engine(str(path))
    cells = [[[] for _ in range(len(xs) - 1)] for __ in range(len(ys) - 1)]
    for box, text, score in result or []:
        if score < 0.45:
            continue
        center_x = sum(point[0] for point in box) / 4
        center_y = sum(point[1] for point in box) / 4
        col = next((idx for idx in range(len(xs) - 1) if xs[idx] <= center_x <= xs[idx + 1]), None)
        row = next((idx for idx in range(len(ys) - 1) if ys[idx] <= center_y <= ys[idx + 1]), None)
        if col is None or row is None:
            continue
        cells[row][col].append((min(point[1] for point in box), min(point[0] for point in box), str(text).strip()))

    table = []
    for row in cells:
        table.append([" ".join(item[2] for item in sorted(cell)).strip() for cell in row])
    return _parse_order_table(table, path.name)


def _parse_order_image_from_ocr_lines(path: Path) -> list[dict[str, Any]]:
    from rapidocr_onnxruntime import RapidOCR

    engine = RapidOCR()
    result, _ = engine(str(path))
    if not result:
        return []
    lines = [str(item[1]).strip() for item in sorted(result, key=lambda item: (min(p[1] for p in item[0]), min(p[0] for p in item[0])))]
    return _parse_order_table([lines], path.name)


def _read_cv_image(path: Path, cv2, np):
    data = np.fromfile(str(path), dtype=np.uint8)
    if data.size == 0:
        return None
    return cv2.imdecode(data, cv2.IMREAD_COLOR)


def _merge_positions(values) -> list[int]:
    values = [int(value) for value in values]
    if not values:
        return []
    groups: list[list[int]] = [[values[0]]]
    for value in values[1:]:
        if value - groups[-1][-1] <= 3:
            groups[-1].append(value)
        else:
            groups.append([value])
    return [round(sum(group) / len(group)) for group in groups]


def _parse_order_table(table: list[list[str]], source_file: str) -> list[dict[str, Any]]:
    headered_rows = _parse_headered_invoice_table(table, source_file)
    if headered_rows:
        return headered_rows
    rows: list[dict[str, Any]] = []
    default_unit = _inline_quantity_unit(table)
    if len(table) == 1 and len(table[0]) >= 3:
        rows.extend(_parse_compact_sequence_cells([_clean_cell(cell) for cell in table[0]], source_file, default_unit))
    for raw in table:
        cells = [_clean_cell(cell) for cell in raw]
        joined = " ".join(cells).upper()
        if not any(re.search(r"\bCOLOR\s*:?\s*\d", cell, flags=re.I) or re.search(r"\b\d+\+?\d*\b", cell) for cell in cells):
            continue
        if "COLOR CODE" in joined and len(cells) >= 5:
            cells = _strip_header_prefixes(cells)
        parsed = (
            _parse_order_cells(cells, source_file)
            or _parse_inline_color_quantity_cells(cells, source_file, default_unit)
            or _parse_compact_style_color_quantity_cells(cells, source_file, default_unit)
        )
        if parsed:
            rows.append(parsed)
    if any(row.get("unit") == "Meter" for row in rows):
        for row in rows:
            quantity_text = str(row.get("quantity_text") or "").upper()
            if not re.search(r"\d\s*(?:Y|YD|YDS)\b", quantity_text):
                row["unit"] = "Meter"
                if row.get("unit_price_usd") is not None:
                    row["price_basis"] = "per_meter"
    return rows


def _parse_headered_invoice_table(table: list[list[str]], source_file: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for header_index, raw_header in enumerate(table):
        header = [_clean_cell(cell).upper() for cell in raw_header]
        if not any("COLOR" in cell for cell in header):
            continue
        if not any("QUANTITY" in cell or "/KG" in cell or "KG" == cell for cell in header):
            continue
        color_idx = _find_header_index(header, ["COLOR"])
        kg_idx = _find_header_index(header, ["QUANTITY/KG", "KG"])
        yard_idx = _find_header_index(header, ["QUANTITY/YARD", "YARD", "YDS"])
        price_idx = _find_header_index(header, ["PRICES/KG", "PRICE/KG", "USD/KG"])
        item_idx = _find_header_index(header, ["ITEM", "FABRIC"])
        if color_idx is None:
            continue
        for raw in table[header_index + 1 :]:
            cells = [_clean_cell(cell) for cell in raw]
            color_cell = cells[color_idx] if color_idx < len(cells) else ""
            color_parts = _parse_color_with_code(color_cell)
            if not color_parts:
                continue
            quantity_cell = cells[kg_idx] if kg_idx is not None and kg_idx < len(cells) else ""
            quantity = _first_float(quantity_cell)
            unit = "KG"
            if quantity is None and yard_idx is not None and yard_idx < len(cells):
                quantity_cell = cells[yard_idx]
                quantity = _first_float(quantity_cell)
                unit = "Yard"
            if quantity is None:
                continue
            item_cell = cells[item_idx] if item_idx is not None and item_idx < len(cells) else " ".join(cells)
            unit_price = _first_float(cells[price_idx]) if price_idx is not None and price_idx < len(cells) else None
            color_name, company_color_code = color_parts
            rows.append(
                {
                    "source_file": source_file,
                    "style": _extract(item_cell, r"ITEM\s*:?\s*([A-Z0-9-]+)") or _extract(item_cell, r"\b([A-Z0-9]{3,})\b"),
                    "color_name": color_name,
                    "company_color_code": company_color_code,
                    "display_color": _company_color_label(color_name, company_color_code),
                    "quantity": quantity,
                    "quantity_text": f"{quantity_cell} {unit}".strip(),
                    "unit": unit,
                    "unit_price_usd": unit_price,
                    "price_basis": "per_kg" if unit_price is not None and unit == "KG" else "",
                    "amount_usd": None,
                }
            )
        if rows:
            return rows
    return rows


def _find_header_index(header: list[str], needles: list[str]) -> int | None:
    for index, cell in enumerate(header):
        compact = re.sub(r"\s+", "", cell.upper())
        for needle in needles:
            if re.sub(r"\s+", "", needle.upper()) in compact:
                return index
    return None


def _parse_compact_sequence_cells(cells: list[str], source_file: str, default_unit: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    index = 0
    while index + 2 < len(cells):
        parsed = _parse_compact_style_color_quantity_cells(cells[index : index + 3], source_file, default_unit)
        if parsed:
            rows.append(parsed)
            index += 3
        else:
            index += 1
    return rows


def _parse_compact_style_color_quantity_cells(cells: list[str], source_file: str, default_unit: str = "Yard") -> dict[str, Any] | None:
    if len(cells) < 3:
        return None
    style = cells[0].strip()
    color_cell = cells[1].strip()
    quantity_cell = cells[2].strip()
    if not style or not color_cell or not re.search(r"\d", quantity_cell):
        return None
    if re.search(r"COLOR|TOTAL|QTY|AMOUNT|PRICE", " ".join(cells), flags=re.I):
        return None
    color_parts = _parse_compact_color_cell(color_cell)
    if not color_parts:
        return None
    quantity = _first_float(quantity_cell)
    if quantity is None:
        return None
    unit = _quantity_unit(quantity_cell, cells)
    if unit == "Meter" and not re.search(r"\d\s*M\b|\(\s*M\s*\)|METER|METRE", " ".join(cells), flags=re.I):
        unit = default_unit
    color_name, company_color_code = color_parts
    unit_price = _price_from_cells(cells, 1)
    amount = _amount_from_cells(cells, 1)
    return {
        "source_file": source_file,
        "style": style,
        "color_name": color_name,
        "company_color_code": company_color_code,
        "display_color": _company_color_label(color_name, company_color_code),
        "quantity": quantity,
        "quantity_text": f"{quantity_cell} {unit}".strip(),
        "unit": unit,
        "unit_price_usd": unit_price,
        "price_basis": "per_meter" if unit_price is not None and unit == "Meter" else ("per_yard" if unit_price is not None and unit == "Yard" else ""),
        "amount_usd": amount,
    }


def _parse_compact_color_cell(value: str) -> tuple[str, str] | None:
    return _parse_color_with_code(value, leading_code=True)


def _parse_color_with_code(value: str, leading_code: bool = False) -> tuple[str, str] | None:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    patterns = []
    if leading_code:
        patterns.append(r"^(?P<code>[0-9O]+(?:\s*[+]\s*[0-9O]+)*)\s*#?\s*(?P<name>.+)$")
    patterns.extend(
        [
            r"^(?P<name>.+?)\s*[-/]\s*#?(?P<code>[0-9O]+(?:\s*[+]\s*[0-9O]+)*)#?$",
            r"^(?P<name>.+?)\s+#?(?P<code>[0-9O]+(?:\s*[+]\s*[0-9O]+)*)#$",
            r"^(?P<name>[A-Z][A-Z /&.,]*?)(?P<code>[0-9O]+(?:\s*[+]\s*[0-9O]+)*)#$",
        ]
    )
    match = next((re.match(pattern, text, flags=re.I) for pattern in patterns if re.match(pattern, text, flags=re.I)), None)
    if not match:
        return None
    raw_code = re.sub(r"\s+", "", match.group("code").upper().replace("O", "0"))
    name = match.group("name").strip(" -/#")
    if not raw_code or not name:
        return None
    return name, _format_company_color_code(raw_code)


def _inline_quantity_unit(table: list[list[str]]) -> str:
    text = " ".join(" ".join(str(cell or "") for cell in row) for row in table).upper()
    if re.search(r"\(\s*Y\s*\)|\bYDS?\b|\bYARD", text):
        return "Yard"
    if re.search(r"\(\s*M\s*\)|\bMETER|\bMETRE", text):
        return "Meter"
    if re.search(r"\bKG\b|\bKGS\b", text):
        return "KG"
    return "Yard"


def _parse_inline_color_quantity_cells(cells: list[str], source_file: str, default_unit: str = "Yard") -> dict[str, Any] | None:
    if len(cells) < 2:
        return None
    color_cell = cells[0].strip()
    if not color_cell or re.search(r"COLOR|TOTAL|QTY|AMOUNT|PRICE", color_cell, flags=re.I):
        return None
    match = re.match(r"^(?P<name>[A-Z][A-Z0-9 /&.,]*?)\s*[-/#]\s*#?(?P<code>[0-9]+(?:\s*[+]\s*[0-9]+)*)#?$", color_cell, flags=re.I)
    if not match:
        return None
    quantity_cell = _inline_quantity_cell(cells)
    quantity = _first_float(quantity_cell)
    if quantity is None:
        return None
    unit = _quantity_unit(quantity_cell, cells)
    if unit == "Meter" and not re.search(r"\d\s*M\b|\(\s*M\s*\)|METER|METRE", " ".join(cells), flags=re.I):
        unit = default_unit
    color_name = re.sub(r"\s+", " ", match.group("name")).strip()
    company_color_code = _format_company_color_code(re.sub(r"\s+", "", match.group("code")))
    unit_price = _price_from_cells(cells, 0)
    amount = _amount_from_cells(cells, 0)
    return {
        "source_file": source_file,
        "style": "",
        "color_name": color_name,
        "company_color_code": company_color_code,
        "display_color": _company_color_label(color_name, company_color_code),
        "quantity": quantity,
        "quantity_text": f"{quantity_cell} {unit}".strip(),
        "unit": unit,
        "unit_price_usd": unit_price,
        "price_basis": "per_meter" if unit_price is not None and unit == "Meter" else ("per_yard" if unit_price is not None and unit == "Yard" else ""),
        "amount_usd": amount,
    }


def _inline_quantity_cell(cells: list[str]) -> str:
    for cell in cells[1:]:
        text = str(cell or "").strip()
        if re.search(r"\d", text) and not re.search(r"COLOR|PRICE|AMOUNT|TOTAL", text, flags=re.I):
            return text
    return ""


def _parse_order_cells(cells: list[str], source_file: str) -> dict[str, Any] | None:
    if len(cells) < 4:
        return None
    style = _strip_header(cells[0], ["STYLE#", "STYLE"])
    color_name = _strip_header(cells[1], ["COLOR NAME", "BODY FABRIC COLOR", "FABRIC COLOR", "COLORNAME"])
    color_code_index = next((index for index, cell in enumerate(cells) if re.search(r"COLOR\s*:?\s*\d", cell, flags=re.I)), -1)
    color_code_cell = cells[color_code_index] if color_code_index >= 0 else ""
    color_code = _extract_color_code(color_code_cell)
    company_color_code = _format_company_color_code(color_code)
    if not color_name or not company_color_code:
        return None
    quantity_cell = _quantity_cell(cells, color_code_index)
    quantity = _first_float(quantity_cell)
    if quantity is None:
        return None
    unit = _quantity_unit(quantity_cell, cells)
    unit_price = _price_from_cells(cells, color_code_index)
    amount = _amount_from_cells(cells, color_code_index)
    return {
        "source_file": source_file,
        "style": style,
        "color_name": color_name,
        "company_color_code": company_color_code,
        "display_color": _company_color_label(color_name, company_color_code),
        "quantity": quantity,
        "quantity_text": quantity_cell,
        "unit": unit,
        "unit_price_usd": unit_price,
        "price_basis": "per_meter" if unit_price is not None and unit == "Meter" else ("per_yard" if unit_price is not None and unit == "Yard" else ""),
        "amount_usd": amount,
    }


def _invoice_rows_from_order_images(rows: list[dict[str, Any]], fabric: FabricRecord | None, fallback: pd.DataFrame) -> pd.DataFrame:
    fallback_price = _fallback_usd_per_kg(fallback)
    quantification = float(fabric.quantification_m_per_kg or 0) if fabric else 0.0
    invoice_rows = []
    for row in rows:
        unit = row.get("unit") or "Yard"
        unit_price = row.get("unit_price_usd")
        usd_per_kg = fallback_price
        if unit_price is not None and quantification:
            if unit == "Meter":
                usd_per_kg = float(unit_price) * quantification
            elif unit == "Yard":
                usd_per_kg = float(unit_price) * quantification * 0.9144
        invoice_rows.append(
            [
                row.get("display_color") or row.get("color_name") or "",
                _company_art_no(row.get("color_name", ""), row.get("company_color_code", "")),
                None,
                row.get("quantity"),
                unit,
                usd_per_kg,
                row.get("style") or "",
                f"客户截图: {row.get('source_file', '')}",
            ]
        )
    return pd.DataFrame(invoice_rows, columns=INVOICE_COLUMNS) if invoice_rows else fallback


def _fallback_usd_per_kg(df: pd.DataFrame) -> float:
    if df is not None and not df.empty and "usd_price_per_kg" in df.columns:
        values = pd.to_numeric(df["usd_price_per_kg"], errors="coerce").dropna()
        if not values.empty:
            return float(values.iloc[0])
    return 0.0


def _strip_header_prefixes(cells: list[str]) -> list[str]:
    return [
        _strip_header(cells[0], ["STYLE#", "STYLE"]),
        _strip_header(cells[1], ["COLOR NAME", "BODY FABRIC COLOR", "COLORNAME"]),
        _strip_header(cells[2], ["COLOR CODE", "BODY FABRIC COLOR CODE", "COLORCODE"]),
        _strip_header(cells[3], ["FABRIC QTY", "TOTAL SMS FABRIC QTY", "(M)"]),
        *cells[4:],
    ]


def _strip_header(value: str, labels: list[str]) -> str:
    text = value.strip()
    compact = re.sub(r"\s+", "", text).upper()
    for label in labels:
        label_compact = re.sub(r"\s+", "", label).upper()
        if compact.startswith(label_compact):
            return text[len(label) :].strip(" #:")
    return text


def _clean_cell(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("：", ":")).strip()


def _extract_color_code(value: str) -> str:
    match = re.search(r"COLOR\s*:?\s*([0-9]+(?:\s*[+]\s*[0-9]+)*)", value, flags=re.I)
    return re.sub(r"\s+", "", match.group(1)) if match else ""


def _format_company_color_code(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    cleaned = text.strip("#").strip()
    return f"{cleaned}#" if cleaned else ""


def _company_color_label(color_name: object, company_color_code: object) -> str:
    name = str(color_name or "").strip()
    code = _format_company_color_code(company_color_code)
    return " ".join(part for part in [name, code] if part)


def _company_art_no(color_name: object, company_color_code: object) -> str:
    name = str(color_name or "").strip()
    code = _format_company_color_code(company_color_code)
    return " - ".join(part for part in [name, code] if part)


def _quantity_cell(cells: list[str], color_code_index: int = -1) -> str:
    search_cells = cells[color_code_index + 1 :] if color_code_index >= 0 else cells
    candidates = [
        cell
        for cell in search_cells
        if "COLOR" not in cell.upper()
        and "$" not in cell
        and not re.search(r"[A-Z]{2,}\d{3,}", cell, flags=re.I)
        and re.search(r"\d+(?:\.\d+)?\s*(?:Y|YD|YDS|M|米)?$", cell, flags=re.I)
    ]
    if candidates:
        return candidates[0]
    return cells[3] if len(cells) > 3 else ""


def _quantity_unit(quantity_cell: str, cells: list[str]) -> str:
    text = " ".join([quantity_cell, *cells]).upper()
    if re.search(r"\d\s*(?:Y|YD|YDS)\b", text):
        return "Yard"
    if "(M)" in text or re.search(r"\d\s*M\b", text):
        return "Meter"
    return "Meter" if "FABRIC QTY" in text else "Yard"


def _price_from_cells(cells: list[str], color_code_index: int = -1) -> float | None:
    search_cells = cells[color_code_index + 1 :] if color_code_index >= 0 else cells
    prices = [_money(cell) for cell in search_cells]
    prices = [price for price in prices if price is not None]
    return prices[0] if prices else None


def _amount_from_cells(cells: list[str], color_code_index: int = -1) -> float | None:
    search_cells = cells[color_code_index + 1 :] if color_code_index >= 0 else cells
    prices = [_money(cell) for cell in search_cells]
    prices = [price for price in prices if price is not None]
    return prices[-1] if len(prices) >= 2 else None


def _money(value: str) -> float | None:
    match = re.search(r"[$S]\s*([0-9][0-9,]*(?:\.\d+)?)", value)
    return float(match.group(1).replace(",", "")) if match else None


def _first_float(value: str) -> float | None:
    match = re.search(r"\d+(?:\.\d+)?", value)
    return float(match.group(0)) if match else None


def _invoice_rows_from_template(template_path: Path | None) -> pd.DataFrame:
    if not template_path or not template_path.exists():
        return pd.DataFrame(columns=INVOICE_COLUMNS)
    try:
        wb_values = load_workbook(template_path, data_only=True)
        ws_values = wb_values.active
        wb_formulas = load_workbook(template_path, data_only=False)
        ws_formulas = wb_formulas.active
    except Exception:
        return pd.DataFrame(columns=INVOICE_COLUMNS)
    rows = []
    for row_no in range(17, 21):
        color = str(ws_values[f"C{row_no}"].value or "").replace("\n", " ").strip()
        kg = ws_values[f"D{row_no}"].value
        price = ws_values[f"F{row_no}"].value
        if not color or kg in (None, "") or price in (None, ""):
            continue
        description = str(ws_formulas[f"A{row_no}"].value or "")
        source_code = _extract(description, r"code:\s*([^\n]+)")
        art_no = _extract(description, r"ART NO\.?:\s*([^\n]+)") or color
        try:
            rows.append([color, art_no.strip(), None, float(kg), "KG", float(price), source_code.strip(), "P.I template"])
        except (TypeError, ValueError):
            continue
    return pd.DataFrame(rows, columns=INVOICE_COLUMNS)


def _roll_rows_from_packing_template(template_path: Path | None) -> pd.DataFrame:
    if not template_path or not template_path.exists():
        return pd.DataFrame(columns=ROLL_COLUMNS)
    try:
        wb = load_workbook(template_path, data_only=True)
        ws = wb.active
    except Exception:
        return pd.DataFrame(columns=ROLL_COLUMNS)
    blocks = [
        {"color": 1, "lot": 2, "net": 4, "price_row": 6},
        {"color": 7, "lot": 8, "net": 10, "price_row": 7},
        {"color": 13, "lot": 14, "net": 16, "price_row": 8},
        {"color": 19, "lot": 20, "net": 22, "price_row": 9},
    ]
    item = str(ws["AB38"].value or "").strip() or _filename_fabric_code(template_path) or ""
    rows = []
    for block in blocks:
        color = str(ws.cell(10, block["color"]).value or ws.cell(block["price_row"], 26).value or "").replace("\n", " ").strip()
        price = ws.cell(block["price_row"], 28).value
        for row_no in range(10, max(ws.max_row, 38) + 1):
            lot = str(ws.cell(row_no, block["lot"]).value or "").strip()
            net = ws.cell(row_no, block["net"]).value
            if lot.casefold() in {"total", "gross weight", "yield"}:
                continue
            if not color or not lot or net in (None, ""):
                continue
            try:
                rows.append([item, color, lot, float(net), float(price or 0)])
            except (TypeError, ValueError):
                continue
    return pd.DataFrame(rows, columns=ROLL_COLUMNS)


def _invoice_template_defaults(template_path: Path | None) -> dict[str, str]:
    if not template_path or not template_path.exists():
        return {}
    wb = load_workbook(template_path, data_only=False)
    ws = wb.active
    row_a17 = str(ws["A17"].value or "")
    row_b17 = str(ws["B17"].value or "")
    return {
        "pi_no": _after(ws["G5"].value, "P/I NO:"),
        "template_date": _after(ws["G6"].value, "DATE:"),
        "buyer": _after(ws["A11"].value, "BUYER:"),
        "buyer_address": str(ws["A12"].value or "").strip(),
        "po_no": _extract(row_a17, r"PO NO\.?:\s*([^\n]+)"),
        "fabric_code": _extract(row_b17, r"ITEM:\s*([^\n]+)") or _filename_fabric_code(template_path),
        "payment_terms": _after(ws["A25"].value, "TERMS OF PAYMENT:"),
        "delivery_time": _after(ws["A27"].value, "DELIVERY TIME:"),
        "port_destination": _after(ws["A29"].value, "PORT OF DESTINATION:"),
        "port_loading": _after(ws["A29"].value, "PORT OF LOADING:").split(".")[0].strip(),
    }


def _packing_template_defaults(template_path: Path | None) -> dict[str, str]:
    if not template_path or not template_path.exists():
        return {}
    wb = load_workbook(template_path, data_only=False)
    ws = wb.active
    header = str(ws["G6"].value or "")
    return {
        "buyer": str(ws["G5"].value or "").strip(),
        "pi_no": _after(ws["P5"].value, "P/I NO:"),
        "template_date": _after(ws["P6"].value, "DATE:"),
        "fabric_code": str(ws["AB38"].value or "").strip() or _extract(header, r"ITEM:\s*([^\n]+)"),
        "po_no": _extract(header, r"PO NO\.?:\s*([^\n]+)"),
    }


def _build_order(defaults: dict[str, str], pl_defaults: dict[str, str], fabric_code: str, invoice_lines: pd.DataFrame) -> OrderInfo:
    total = float(invoice_lines["amount_usd"].sum()) if not invoice_lines.empty else 0.0
    return OrderInfo(
        po_no=defaults.get("po_no") or pl_defaults.get("po_no") or "",
        art_no=fabric_code,
        fabric_code=fabric_code,
        buyer=defaults.get("buyer") or pl_defaults.get("buyer") or "",
        buyer_address=defaults.get("buyer_address") or "",
        pi_no=suggest_doc_number(today=_today()),
        ci_no=suggest_doc_number(today=_today()),
        order_date=_today().strftime("%Y-%m-%d"),
        advance_payment_usd=round(total * 0.3, 2),
        output_dir=OUTPUT_DIR,
        payment_terms=defaults.get("payment_terms") or "30% deposit and 70% before shipment",
        delivery_time=defaults.get("delivery_time") or "",
        port_destination=defaults.get("port_destination") or "",
        port_loading=defaults.get("port_loading") or "GUANGZHOU,CHINA",
    )


def _build_issues(
    order: OrderInfo,
    invoice_lines: pd.DataFrame,
    computed_rolls: pd.DataFrame,
    packing_summary: pd.DataFrame,
    fabric: FabricRecord | None,
    paths: dict[str, Path],
) -> list[dict[str, Any]]:
    total = float(invoice_lines["amount_usd"].sum()) if not invoice_lines.empty else 0.0
    invoice_kg = float(invoice_lines["total_net_weight_kg"].sum()) if not invoice_lines.empty else 0.0
    packing_kg = float(computed_rolls["net_weight_kg"].sum()) if not computed_rolls.empty else 0.0
    return [
        _issue("buyer_address", "买方地址 Buyer’s address", order.buyer_address, "从 P.I 模板识别，生成前建议确认是否完整。", bool(order.buyer_address)),
        _issue("payment_terms", "付款方式 Terms of payment", order.payment_terms, "会写入 P.I 的 TERMS OF PAYMENT。", bool(order.payment_terms)),
        _issue("delivery_time", "交货期 Delivery Time", order.delivery_time, "会写入 P.I 的 DELIVERY TIME。", bool(order.delivery_time)),
        _issue("port_destination", "目的港 Port of Destination", order.port_destination, "会写入 P.I 的 PORT OF DESTINATION。", bool(order.port_destination)),
        _issue("deposit", "定金 30%", f"${total * 0.3:,.2f}", "Total Amount × 30%，系统自动计算。", total > 0, auto=True),
        _issue("weight_match", "P.I / P.L 净重匹配", f"P.I {invoice_kg:,.2f} KG / P.L {packing_kg:,.2f} KG", "两份单据总净重应一致。", abs(invoice_kg - packing_kg) < 0.05 and invoice_kg > 0, auto=True),
        _issue("templates", "模板与面料库", _template_status(paths, fabric, packing_summary), "缺少模板或面料数据会阻止生成。", _can_generate(invoice_lines, packing_summary, fabric, paths), auto=True),
    ]


def _issue(key: str, title: str, value: str, detail: str, passed: bool, auto: bool = False) -> dict[str, Any]:
    return {
        "id": key,
        "title": title,
        "value": value or "",
        "detail": detail,
        "severity": "pass" if passed else "attention",
        "confidence": "已校验" if passed and auto else ("已识别" if passed else "需要确认"),
        "editable": not auto,
    }


def _compute_invoice(fabric: FabricRecord | None, invoice_input: pd.DataFrame) -> pd.DataFrame:
    if not fabric:
        return pd.DataFrame()
    return calculate_invoice_lines(invoice_input, float(fabric.quantification_m_per_kg or 0))


def _compute_packing(fabric: FabricRecord | None, roll_input: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    if not fabric or not fabric.quantification_m_per_kg:
        return pd.DataFrame(), pd.DataFrame()
    computed_rolls, summary = calculate_rolls(roll_input, float(fabric.quantification_m_per_kg))
    if fabric.tube_plus_allowance_kg_per_roll is not None:
        summary = apply_gross_weight(summary, float(fabric.tube_plus_allowance_kg_per_roll))
    return computed_rolls, summary


def _load_fabric_df(path: Path | None) -> pd.DataFrame:
    if path and path.exists():
        return load_fabric_master(path)
    return pd.DataFrame()


def _fabric_from_template(fabric_code: str, defaults: dict[str, str], pl_defaults: dict[str, str]) -> FabricRecord:
    return FabricRecord(fabric_code=fabric_code or "UNKNOWN", raw={"source": "template fallback"})


def _file_status(paths: dict[str, Path], order_images: list[Path] | None = None) -> list[dict[str, str]]:
    labels = {"pi_template": "P.I 模板", "pl_template": "P.L 模板", "fabric_db": "面料库"}
    files = [
        {"kind": key, "label": label, "name": paths[key].name if key in paths else "未选择", "status": "已识别" if key in paths else "缺少"}
        for key, label in labels.items()
    ]
    if order_images:
        files.append(
            {
                "kind": "order_image",
                "label": "客户截图",
                "name": f"{len(order_images)} 张颜色/价格截图",
                "status": "已解析",
            }
        )
    return files


def _fabric_payload(fabric: FabricRecord | None) -> dict[str, Any] | None:
    if not fabric:
        return None
    return {
        "fabric_code": fabric.fabric_code,
        "name_cn": fabric.fabric_name_cn,
        "name_en": fabric.fabric_name_en,
        "composition": fabric.composition_en or fabric.composition_cn,
        "weight": fabric.weight_en or fabric.weight_cn,
        "width": fabric.width_en or fabric.width_cn,
        "quantification_m_per_kg": fabric.quantification_m_per_kg,
        "tube_plus_allowance_kg_per_roll": fabric.tube_plus_allowance_kg_per_roll,
        "grade": fabric.quality_grade_en or fabric.quality_grade_cn,
    }


def _totals(invoice_lines: pd.DataFrame, computed_rolls: pd.DataFrame, packing_summary: pd.DataFrame) -> dict[str, Any]:
    total = float(invoice_lines["amount_usd"].sum()) if not invoice_lines.empty else 0.0
    return {
        "total_amount": total,
        "deposit": total * 0.3,
        "balance": total * 0.7,
        "invoice_kg": float(invoice_lines["total_net_weight_kg"].sum()) if not invoice_lines.empty else 0.0,
        "packing_kg": float(computed_rolls["net_weight_kg"].sum()) if not computed_rolls.empty else 0.0,
        "rolls": int(len(computed_rolls)) if not computed_rolls.empty else 0,
        "groups": int(len(packing_summary)) if not packing_summary.empty else 0,
    }


def _can_generate(invoice_lines: pd.DataFrame, packing_summary: pd.DataFrame, fabric: FabricRecord | None, paths: dict[str, Path]) -> bool:
    return bool(
        fabric
        and paths.get("pi_template")
        and paths.get("pl_template")
        and not invoice_lines.empty
        and not packing_summary.empty
        and len(packing_summary) <= 4
    )


def _template_status(paths: dict[str, Path], fabric: FabricRecord | None, summary: pd.DataFrame) -> str:
    parts = []
    parts.append("P.I ✓" if paths.get("pi_template") else "P.I 缺少")
    parts.append("P.L ✓" if paths.get("pl_template") else "P.L 缺少")
    parts.append("面料 ✓" if fabric else "面料缺少")
    if not summary.empty and len(summary) > 4:
        parts.append("P.L 分组超过 4 组")
    return " / ".join(parts)


def _records(df: pd.DataFrame) -> list[dict[str, Any]]:
    if df is None or df.empty:
        return []
    clean = df.where(pd.notnull(df), None)
    return clean.to_dict(orient="records")


def _download_payload(session_id: str, kind: str, path: Path) -> dict[str, str]:
    return {"name": path.name, "url": f"/api/download/{session_id}/{kind}", "path": str(path)}


def _write_manifest(session_path: Path, manifest: dict[str, Any]) -> None:
    (session_path / "manifest.json").write_text(json.dumps(manifest, indent=2, ensure_ascii=False, default=str), encoding="utf-8")


def _read_manifest(session_path: Path) -> dict[str, Any]:
    path = session_path / "manifest.json"
    if not path.exists():
        raise FileNotFoundError("会话不存在，请重新分析文件。")
    return json.loads(path.read_text(encoding="utf-8"))


def _order_kwargs(data: dict[str, Any]) -> dict[str, Any]:
    allowed = OrderInfo.__dataclass_fields__.keys()
    values = {key: data.get(key) for key in allowed if key in data}
    values["output_dir"] = Path(values.get("output_dir") or OUTPUT_DIR)
    return values


def _extract(value: str, pattern: str) -> str:
    match = re.search(pattern, str(value or ""), flags=re.I)
    return match.group(1).strip() if match else ""


def _after(value: object, label: str) -> str:
    text = str(value or "").strip()
    index = text.upper().find(label.upper())
    if index < 0:
        return text
    return text[index + len(label) :].strip()


def _filename_fabric_code(path: Path | None) -> str:
    if not path:
        return ""
    match = re.search(r"(?:^|[-_ ])(\d{4,6}[A-Z]?)(?:[-_ .]|$)", path.name, flags=re.I)
    return match.group(1) if match else ""


def _safe_filename(name: str) -> str:
    cleaned = re.sub(r'[<>:"/\\\\|?*]+', "_", name).strip()
    return cleaned or f"upload-{uuid.uuid4().hex}"


def _today():
    return datetime.now(tz=ZoneInfo("Asia/Shanghai")).date()
