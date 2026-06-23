from __future__ import annotations

import json
import re
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from .models import FabricRecord, GeneratedFile, GeneratedInvoiceFiles, OrderInfo
from .numbering import commit_doc_number


BLOCKS = [
    {"color": 1, "lot": 2, "roll": 3, "net": 4, "meter": 5, "yard": 6, "gross": 4},
    {"color": 7, "lot": 8, "roll": 9, "net": 10, "meter": 11, "yard": 12, "gross": 10},
    {"color": 13, "lot": 14, "roll": 15, "net": 16, "meter": 17, "yard": 18, "gross": 16},
    {"color": 19, "lot": 20, "roll": 21, "net": 22, "meter": 23, "yard": 24, "gross": 22},
]
DATA_START_ROW = 10
DATA_END_ROW = 37
TOTAL_ROW = 38
GROSS_ROW = 40
REVIEW_FILL = PatternFill("solid", fgColor="FFF2CC")


@dataclass(frozen=True)
class PackingLayout:
    data_start_row: int
    data_end_row: int
    total_row: int
    yield_row: int
    gross_row: int
    is_yield_template: bool


def write_packing_list(
    template_path: str | Path,
    order: OrderInfo,
    fabric: FabricRecord,
    rolls_df: pd.DataFrame,
    summary_df: pd.DataFrame,
) -> GeneratedFile:
    template_path = Path(template_path)
    output_dir = Path(order.output_dir)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    run_dir = output_dir / stamp
    run_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(template_path)
    ws = wb.active

    _write_header(ws, order, fabric)
    layout = _prepare_packing_layout(ws, summary_df)
    _clear_roll_blocks(ws, layout)
    _write_roll_blocks(ws, rolls_df, summary_df, fabric, layout)
    _write_summary_area(ws, order, summary_df, layout)
    _write_gross_area(ws, order, summary_df, fabric, layout)

    filename = f"Packing List-{_safe(order.po_no)}-{_safe(order.art_no or order.fabric_code)}-{order.order_date.replace('-', '')}.xlsx"
    output_path = run_dir / filename
    _force_recalculate_on_open(wb)
    wb.save(output_path)

    log = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "template": str(template_path),
        "output": str(output_path),
        "price_source": "USD/KG from PI/CI or user input",
        "order": {
            "po_no": order.po_no,
            "art_no": order.art_no,
            "fabric_code": order.fabric_code,
            "buyer": order.buyer,
            "buyer_address": order.buyer_address,
            "pi_no": order.pi_no,
            "ci_no": order.ci_no,
            "order_date": order.order_date,
            "advance_payment_usd": order.advance_payment_usd,
            "note": order.note,
        },
        "fabric": fabric.raw,
        "rolls": rolls_df.to_dict(orient="records"),
        "summary": summary_df.to_dict(orient="records"),
    }
    log_path = run_dir / f"{output_path.stem}.json"
    log_path.write_text(json.dumps(log, indent=2, ensure_ascii=False), encoding="utf-8")
    commit_doc_number(order.pi_no)
    return GeneratedFile(path=output_path, log_path=log_path)


def write_invoice_pair(
    template_path: str | Path,
    order: OrderInfo,
    fabric: FabricRecord,
    invoice_df: pd.DataFrame,
    exchange_rate: float | None = None,
    exchange_source: str = "",
    quantity_unit: str = "Yard",
) -> GeneratedInvoiceFiles:
    template_path = Path(template_path)
    output_dir = Path(order.output_dir)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    run_dir = output_dir / stamp
    run_dir.mkdir(parents=True, exist_ok=True)

    pi_path = _write_invoice(
        template_path=template_path,
        run_dir=run_dir,
        order=order,
        fabric=fabric,
        invoice_df=invoice_df,
        title="PROFORMA INVOICE",
        doc_label="P/I NO",
        doc_no=order.pi_no,
        prefix="PI",
        quantity_unit=quantity_unit,
    )
    ci_path = _write_invoice(
        template_path=template_path,
        run_dir=run_dir,
        order=order,
        fabric=fabric,
        invoice_df=invoice_df,
        title="COMMERCIAL INVOICE",
        doc_label="C/I NO",
        doc_no=order.ci_no,
        prefix="CI",
        quantity_unit=quantity_unit,
    )

    log = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "template": str(template_path),
        "pi_output": str(pi_path),
        "ci_output": str(ci_path),
        "exchange_rate_usd_cny": exchange_rate,
        "exchange_source": exchange_source,
        "order": {
            "po_no": order.po_no,
            "art_no": order.art_no,
            "fabric_code": order.fabric_code,
            "buyer": order.buyer,
            "buyer_address": order.buyer_address,
            "pi_no": order.pi_no,
            "ci_no": order.ci_no,
            "order_date": order.order_date,
            "advance_payment_usd": order.advance_payment_usd,
            "note": order.note,
        },
        "fabric": fabric.raw,
        "quantity_unit": quantity_unit,
        "invoice_lines": invoice_df.to_dict(orient="records"),
    }
    log_path = run_dir / f"PI-CI-{_safe(order.po_no)}-{_safe(order.art_no or order.fabric_code)}-{order.order_date.replace('-', '')}.json"
    log_path.write_text(json.dumps(log, indent=2, ensure_ascii=False), encoding="utf-8")
    commit_doc_number(order.pi_no)
    commit_doc_number(order.ci_no)
    return GeneratedInvoiceFiles(pi_path=pi_path, ci_path=ci_path, log_path=log_path)


def write_proforma_invoice(
    template_path: str | Path,
    order: OrderInfo,
    fabric: FabricRecord,
    invoice_df: pd.DataFrame,
    quantity_unit: str = "Yard",
) -> GeneratedFile:
    template_path = Path(template_path)
    output_dir = Path(order.output_dir)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    run_dir = output_dir / stamp
    run_dir.mkdir(parents=True, exist_ok=True)

    pi_path = _write_invoice(
        template_path=template_path,
        run_dir=run_dir,
        order=order,
        fabric=fabric,
        invoice_df=invoice_df,
        title="PROFORMA INVOICE",
        doc_label="P/I NO",
        doc_no=order.pi_no,
        prefix="PI",
        quantity_unit=quantity_unit,
    )
    log = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "template": str(template_path),
        "output": str(pi_path),
        "order": {
            "po_no": order.po_no,
            "art_no": order.art_no,
            "fabric_code": order.fabric_code,
            "buyer": order.buyer,
            "buyer_address": order.buyer_address,
            "pi_no": order.pi_no,
            "order_date": order.order_date,
            "advance_payment_usd": order.advance_payment_usd,
        },
        "fabric": fabric.raw,
        "quantity_unit": quantity_unit,
        "invoice_lines": invoice_df.to_dict(orient="records"),
    }
    log_path = run_dir / f"PI-{_safe(order.po_no)}-{_safe(order.art_no or order.fabric_code)}-{order.order_date.replace('-', '')}.json"
    log_path.write_text(json.dumps(log, indent=2, ensure_ascii=False), encoding="utf-8")
    commit_doc_number(order.pi_no)
    return GeneratedFile(path=pi_path, log_path=log_path)


def _write_invoice(
    template_path: Path,
    run_dir: Path,
    order: OrderInfo,
    fabric: FabricRecord,
    invoice_df: pd.DataFrame,
    title: str,
    doc_label: str,
    doc_no: str,
    prefix: str,
    quantity_unit: str,
) -> Path:
    wb = load_workbook(template_path)
    ws = wb.active
    if _is_pi_yard_price_template(ws):
        _write_pi_yard_price_sheet(ws, order, fabric, invoice_df, title, doc_label, doc_no)
        filename = f"{prefix}-{_safe(order.po_no)}-{_safe(order.art_no or order.fabric_code)}-{order.order_date.replace('-', '')}.xlsx"
        output_path = run_dir / filename
        wb.save(output_path)
        return output_path

    _set(ws, "A4", title)
    _set(ws, "F5", f"{doc_label}:{doc_no}")
    _set(ws, "F6", f"DATE: {order.order_date}")
    _set(ws, "A11", f"BUYER: {order.buyer}")
    _set(ws, "A12", order.buyer_address)
    price_width = ws.column_dimensions["E"].width
    amount_width = ws.column_dimensions["F"].width
    _set(ws, "D16", f"QUANTITY/{quantity_unit.upper()}")
    _set(ws, "E16", "QUANTITY/KG")
    _set(ws, "F16", "PRICES/KG\n(USD)FOB")
    _set(ws, "G16", "TOTAL AMOUNT:")
    ws.column_dimensions["G"].hidden = False
    ws.column_dimensions["F"].width = price_width
    ws.column_dimensions["G"].width = amount_width
    _copy_cell_format(ws, "F16", "G16")

    layout = _prepare_invoice_detail_layout(ws, len(invoice_df))
    for row in range(17, layout["data_end"] + 1):
        for col in range(1, 8):
            _safe_set(ws, row, col, None)
        _copy_cell_format(ws, f"F{row}", f"G{row}")

    description = _invoice_description(order, fabric)
    for offset, item in enumerate(invoice_df.itertuples(index=False), start=17):
        color = str(item.color)
        _safe_set(ws, offset, 1, _invoice_po_line(order, item))
        _safe_set(ws, offset, 2, description)
        _safe_set(ws, offset, 3, color)
        _safe_set(ws, offset, 4, round(_invoice_quantity(item, quantity_unit), 2))
        _safe_set(ws, offset, 5, round(float(item.total_net_weight_kg), 4))
        _safe_set(ws, offset, 6, round(float(item.usd_price_per_kg), 4))
        _safe_set(ws, offset, 7, f"=E{offset}*F{offset}")

    _set(ws, f"D{layout['total_row']}", float(order.advance_payment_usd or 0))
    _set(ws, f"C{layout['balance_row']}", f"=SUM(G17:G{layout['data_end']})-D{layout['total_row']}")
    _write_invoice_review_text(ws, order, layout)
    _highlight_invoice_review_cells(ws, use_yard_price=False, layout=layout)

    filename = f"{prefix}-{_safe(order.po_no)}-{_safe(order.art_no or order.fabric_code)}-{order.order_date.replace('-', '')}.xlsx"
    output_path = run_dir / filename
    _force_recalculate_on_open(wb)
    wb.save(output_path)
    return output_path


def _is_pi_yard_price_template(ws) -> bool:
    header = " ".join(str(ws[cell].value or "") for cell in ["D16", "E16", "F16", "G16", "H16"]).upper()
    return "PRICES/YDS" in header or "QUANTITY/YARD" in header


def _write_pi_yard_price_sheet(
    ws,
    order: OrderInfo,
    fabric: FabricRecord,
    invoice_df: pd.DataFrame,
    title: str,
    doc_label: str,
    doc_no: str,
) -> None:
    _set(ws, "A4", title)
    _set(ws, "G5", f"{doc_label}:{doc_no}")
    _set(ws, "G6", f"DATE: {order.order_date.replace('-', '/')}")
    _set(ws, "A11", f"BUYER: {order.buyer}")
    _set(ws, "A12", order.buyer_address)
    _set(ws, "D16", "QUANTITY/KG")
    _set(ws, "E16", "QUANTITY/Yard")
    _set(ws, "F16", "PRICES/KG\n(USD)FOB")
    _set(ws, "G16", "TOTAL AMOUNT:")
    _set(ws, "H16", "PRICES/YDS\n(USD)FOB")

    layout = _prepare_invoice_detail_layout(ws, len(invoice_df))
    for row in range(17, layout["data_end"] + 1):
        for col in range(1, 9):
            _safe_set(ws, row, col, None)

    description = _invoice_description(order, fabric)
    quantification = _num(fabric.quantification_m_per_kg)
    for row_no, item in enumerate(invoice_df.itertuples(index=False), start=17):
        _safe_set(ws, row_no, 1, _invoice_po_line(order, item))
        _safe_set(ws, row_no, 2, description)
        _safe_set(ws, row_no, 3, str(item.color))
        _safe_set(ws, row_no, 4, round(float(item.total_net_weight_kg), 4))
        _safe_set(ws, row_no, 5, f"=D{row_no}*{quantification}*1.0936")
        _safe_set(ws, row_no, 6, round(float(item.usd_price_per_kg), 4))
        _safe_set(ws, row_no, 7, f"=D{row_no}*F{row_no}")
        _safe_set(ws, row_no, 8, f"=G{row_no}/E{row_no}")

    _set(ws, f"C{layout['total_row']}", f"=SUM(G17:G{layout['data_end']})")
    _set(ws, f"D{layout['deposit_row']}", float(order.advance_payment_usd or 0))
    _set(ws, f"C{layout['balance_row']}", f"=C{layout['total_row']}-D{layout['deposit_row']}")
    _write_invoice_review_text(ws, order, layout)
    _highlight_invoice_review_cells(ws, use_yard_price=True, layout=layout)


def _prepare_invoice_detail_layout(ws, detail_count: int) -> dict[str, int]:
    visible_count = max(4, int(detail_count or 0))
    extra_rows = max(0, visible_count - 4)
    if extra_rows:
        ws.insert_rows(21, extra_rows)
        _unmerge_rows(ws, 21, 20 + extra_rows)
        _copy_row_format(ws, 20, 21, 20 + extra_rows)
    total_row = 21 + extra_rows
    return {
        "data_end": 16 + visible_count,
        "total_row": total_row,
        "balance_row": total_row + 1,
        "deposit_row": total_row + 2,
        "terms_row": total_row + 4,
        "delivery_row": total_row + 6,
        "port_row": total_row + 8,
    }


def _unmerge_rows(ws, start_row: int, end_row: int) -> None:
    ranges = list(ws.merged_cells.ranges)
    for merged_range in ranges:
        if start_row <= merged_range.min_row and merged_range.max_row <= end_row:
            try:
                ws.unmerge_cells(str(merged_range))
            except KeyError:
                try:
                    ws.merged_cells.ranges.remove(merged_range)
                except KeyError:
                    pass
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    if isinstance(ws.cell(row, col), MergedCell):
                        ws._cells[(row, col)] = Cell(ws, row=row, column=col)


def _highlight_invoice_review_cells(ws, use_yard_price: bool, layout: dict[str, int] | None = None) -> None:
    layout = layout or {
        "total_row": 21,
        "balance_row": 22,
        "deposit_row": 23,
        "terms_row": 25,
        "delivery_row": 27,
        "port_row": 29,
    }
    cells = [
        "G5",
        "G6",
        "A11",
        "A12",
        f"C{layout['total_row']}",
        f"C{layout['balance_row']}",
        f"D{layout['deposit_row']}",
        f"A{layout['deposit_row']}",
        f"A{layout['terms_row']}",
        f"A{layout['delivery_row']}",
        f"A{layout['port_row']}",
    ]
    if not use_yard_price:
        cells.extend(["F5", "F6"])
    for coordinate in cells:
        _fill_if_writable(ws, coordinate, REVIEW_FILL)


def _write_invoice_review_text(ws, order: OrderInfo, layout: dict[str, int] | None = None) -> None:
    layout = layout or {"deposit_row": 23, "terms_row": 25, "delivery_row": 27, "port_row": 29}
    if order.advance_payment_usd:
        _set(ws, f"A{layout['deposit_row']}", "Attention：A deposit of 30% of the total cost is requested.")
    if order.payment_terms:
        _set(ws, f"A{layout['terms_row']}", f"1.TERMS OF PAYMENT:{order.payment_terms}")
    if order.delivery_time:
        _set(ws, f"A{layout['delivery_row']}", f"3. DELIVERY TIME: {order.delivery_time}")
    if order.port_destination:
        loading = order.port_loading or "GUANGZHOU,CHINA"
        _set(
            ws,
            f"A{layout['port_row']}",
            f"5.PORT OF LOADING: {loading}.                  PORT OF DESTINATION: {order.port_destination}",
        )


def _invoice_quantity(item, quantity_unit: str) -> float:
    if quantity_unit == "KG":
        return float(item.total_net_weight_kg)
    if quantity_unit == "Meter":
        return float(item.total_meter)
    return float(item.total_yard)


def _invoice_po_line(order: OrderInfo, item) -> str:
    source_code = str(getattr(item, "source_fabric_code", "") or order.fabric_code).strip()
    color = str(getattr(item, "color", "")).strip()
    art_no = str(getattr(item, "art_no", "") or color).strip()
    return (
        f"PO NO.: {order.po_no}\n"
        f"code: {source_code}\n"
        f"ART NO.:{art_no}\n"
        "Knit fabric"
    )


def _invoice_description(order: OrderInfo, fabric: FabricRecord) -> str:
    return (
        f"ITEM: {order.art_no or fabric.fabric_code}\n"
        f"COMP:  {fabric.composition_en or fabric.composition_cn}\n"
        f"WEIGHT :{fabric.weight_en or fabric.weight_cn}        WIDTH :{fabric.width_en or fabric.width_cn}\n"
        "FINISH:\n"
        "COUNTRY OF ORIGIN:CHINA"
    )


def _write_header(ws, order: OrderInfo, fabric: FabricRecord) -> None:
    _set(ws, "G5", order.buyer)
    _set(ws, "P5", f"P/I NO:{order.pi_no}")
    _set(ws, "P6", f"DATE: {order.order_date.replace('-', '/')}")
    _set(ws, "G6", fabric.packing_description(order.art_no))
    for coordinate in ["G5", "G6", "P5", "P6"]:
        _fill_if_writable(ws, coordinate, REVIEW_FILL)


def _prepare_packing_layout(ws, summary_df: pd.DataFrame) -> PackingLayout:
    max_rolls = 0
    if not summary_df.empty and "rolls" in summary_df.columns:
        max_rolls = int(pd.to_numeric(summary_df["rolls"], errors="coerce").fillna(0).max())
    required_capacity = max(DATA_END_ROW - DATA_START_ROW + 1, max_rolls)
    extra_rows = required_capacity - (DATA_END_ROW - DATA_START_ROW + 1)
    is_yield_template = _is_packing_yield_template(ws)
    if extra_rows > 0:
        ws.insert_rows(TOTAL_ROW, extra_rows)
        _copy_row_format(ws, DATA_END_ROW, TOTAL_ROW, TOTAL_ROW + extra_rows - 1)
    data_end_row = DATA_START_ROW + required_capacity - 1
    total_row = data_end_row + 1
    return PackingLayout(
        data_start_row=DATA_START_ROW,
        data_end_row=data_end_row,
        total_row=total_row,
        yield_row=total_row + 1,
        gross_row=total_row + 2,
        is_yield_template=is_yield_template,
    )


def _clear_roll_blocks(ws, layout: PackingLayout) -> None:
    for block in BLOCKS:
        _safe_set(ws, layout.data_start_row, block["color"], None)
        for row in range(layout.data_start_row, layout.data_end_row + 1):
            for key in ["lot", "roll", "net", "meter", "yard"]:
                _safe_set(ws, row, block[key], None)
        for key in ["roll", "net", "meter", "yard"]:
            _safe_set(ws, layout.total_row, block[key], None)
        _safe_set(ws, layout.gross_row, block["gross"], None)

    for row in range(6, 10):
        for col in range(26, 31):
            _safe_set(ws, row, col, None)
    for cell in [
        f"AB{layout.total_row}",
        f"AC{layout.total_row}",
        f"AB{layout.yield_row}",
        f"AC{layout.yield_row}",
        f"AD{layout.yield_row}",
        f"AB{layout.gross_row}",
        f"AC{layout.gross_row}",
        f"AD{layout.gross_row}",
        f"AE{layout.yield_row}",
    ]:
        _clear_coordinate(ws, cell)


def _write_roll_blocks(ws, rolls_df: pd.DataFrame, summary_df: pd.DataFrame, fabric: FabricRecord, layout: PackingLayout) -> None:
    quantification = _num(fabric.quantification_m_per_kg)
    if layout.is_yield_template:
        _set(ws, f"B{layout.yield_row}", float(fabric.quantification_m_per_kg or 0))
        quantification_ref = f"$B${layout.yield_row}"
    else:
        quantification_ref = quantification
    yard_factor = 1 / 0.9144
    groups = list(summary_df[["item", "color", "lot"]].itertuples(index=False, name=None))
    for index, (item, color, lot) in enumerate(groups):
        block = BLOCKS[index]
        group = rolls_df[
            (rolls_df["item"].fillna("").astype(str) == str(item))
            & (rolls_df["color"].astype(str) == str(color))
            & (rolls_df["lot"].astype(str) == str(lot))
        ].copy()
        color_text = str(color)
        _safe_set(ws, layout.data_start_row, block["color"], color_text)

        for offset, row_data in enumerate(group.itertuples(index=False)):
            row = layout.data_start_row + offset
            _safe_set(ws, row, block["lot"], str(lot))
            _safe_set(ws, row, block["roll"], int(getattr(row_data, "roll")))
            _safe_set(ws, row, block["net"], float(getattr(row_data, "net_weight_kg")))
            net_cell = f"{get_column_letter(block['net'])}{row}"
            meter_cell = f"{get_column_letter(block['meter'])}{row}"
            _safe_set(ws, row, block["meter"], f"={net_cell}*{quantification_ref}")
            _safe_set(ws, row, block["yard"], f"={meter_cell}*{_num(yard_factor)}")

        roll_col = get_column_letter(block["roll"])
        net_col = get_column_letter(block["net"])
        meter_col = get_column_letter(block["meter"])
        yard_col = get_column_letter(block["yard"])
        _safe_set(ws, layout.total_row, block["roll"], f"=COUNT({roll_col}{layout.data_start_row}:{roll_col}{layout.data_end_row})")
        _safe_set(ws, layout.total_row, block["net"], f"=SUM({net_col}{layout.data_start_row}:{net_col}{layout.data_end_row})")
        _safe_set(ws, layout.total_row, block["meter"], f"=SUM({meter_col}{layout.data_start_row}:{meter_col}{layout.data_end_row})")
        _safe_set(ws, layout.total_row, block["yard"], f"=SUM({yard_col}{layout.data_start_row}:{yard_col}{layout.data_end_row})")


def _write_summary_area(ws, order: OrderInfo, summary_df: pd.DataFrame, layout: PackingLayout) -> None:
    for i, row in enumerate(summary_df.itertuples(index=False), start=6):
        block_index = i - 6
        _set(ws, f"Z{i}", str(row.color))
        if layout.is_yield_template and block_index < len(BLOCKS):
            net_col = get_column_letter(BLOCKS[block_index]["net"])
            _set(ws, f"AA{i}", f"={net_col}{layout.total_row}")
        else:
            _set(ws, f"AA{i}", round(float(row.total_net_weight_kg), 2))
        _set(ws, f"AB{i}", round(float(row.usd_price_per_kg), 4))
        _set(ws, f"AC{i}", f"=AA{i}*AB{i}")
    _set(ws, "AD6", float(order.advance_payment_usd or 0))
    _set(ws, "AE6", "=SUM(AC6:AC9)-AD6")
    for row in range(5, 10):
        for col in range(26, 31):
            _fill_if_writable(ws, f"{get_column_letter(col)}{row}", REVIEW_FILL)


def _write_gross_area(ws, order: OrderInfo, summary_df: pd.DataFrame, fabric: FabricRecord, layout: PackingLayout) -> None:
    allowance = _num(fabric.tube_plus_allowance_kg_per_roll)
    if layout.is_yield_template:
        _set(ws, f"B{layout.yield_row}", float(fabric.quantification_m_per_kg or 0))
        _set(ws, f"E{layout.yield_row}", float(fabric.tube_plus_allowance_kg_per_roll or 0))
        _set(ws, f"AB{layout.total_row}", order.art_no or order.fabric_code)
        _fill_if_writable(ws, f"B{layout.yield_row}", REVIEW_FILL)
        _fill_if_writable(ws, f"E{layout.yield_row}", REVIEW_FILL)
    else:
        _set(ws, "AB39", order.art_no or order.fabric_code)
        _set(ws, "AC39", order.fabric_code)
    for block in BLOCKS[: len(summary_df)]:
        net_cell = f"{get_column_letter(block['net'])}{layout.total_row}"
        roll_cell = f"{get_column_letter(block['roll'])}{layout.total_row}"
        allowance_ref = f"$E${layout.yield_row}" if layout.is_yield_template else allowance
        _safe_set(ws, layout.gross_row, block["gross"], f"={net_cell}+{roll_cell}*{allowance_ref}")
    gross_cells = ",".join(f"{get_column_letter(block['gross'])}{layout.gross_row}" for block in BLOCKS[: len(summary_df)])
    if layout.is_yield_template:
        _set(ws, f"AB{layout.yield_row}", f"=SUM({gross_cells})" if gross_cells else 0)
        _set(ws, f"AC{layout.yield_row}", f"=SUM(AB{layout.yield_row})")
        _set(ws, f"AE{layout.yield_row}", order.note or None)
        for coordinate in [f"AB{layout.yield_row}", f"AC{layout.yield_row}", f"AE{layout.yield_row}"]:
            _fill_if_writable(ws, coordinate, REVIEW_FILL)
    else:
        _set(ws, f"AB{layout.gross_row}", f"=SUM({gross_cells})" if gross_cells else 0)
        _set(ws, f"AD{layout.gross_row}", f"=AB{layout.gross_row}")
        _set(ws, f"AE{layout.gross_row}", order.note or None)


def _is_packing_yield_template(ws) -> bool:
    return "YIELD" in str(ws["A39"].value or "").upper() or str(ws["B39"].value or "").strip() == "3.25"


def _set(ws, coordinate: str, value) -> None:
    cell = ws[coordinate]
    if isinstance(cell, MergedCell):
        raise ValueError(f"Cannot write to merged child cell {coordinate}")
    cell.value = value


def _clear_coordinate(ws, coordinate: str) -> None:
    cell = ws[coordinate]
    if isinstance(cell, MergedCell):
        return
    cell.value = None


def _safe_set(ws, row: int, col: int, value) -> None:
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def _copy_cell_format(ws, source: str, target: str) -> None:
    source_cell = ws[source]
    target_cell = ws[target]
    if isinstance(source_cell, MergedCell) or isinstance(target_cell, MergedCell):
        return
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)
    if source_cell.border:
        target_cell.border = copy(source_cell.border)
    if source_cell.fill:
        target_cell.fill = copy(source_cell.fill)
    if source_cell.font:
        target_cell.font = copy(source_cell.font)


def _copy_row_format(ws, source_row: int, start_row: int, end_row: int) -> None:
    for row in range(start_row, end_row + 1):
        ws.row_dimensions[row].height = ws.row_dimensions[source_row].height
        for col in range(1, ws.max_column + 1):
            source_cell = ws.cell(source_row, col)
            target_cell = ws.cell(row, col)
            if isinstance(target_cell, MergedCell):
                continue
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            target_cell.number_format = source_cell.number_format
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.font = copy(source_cell.font)


def _fill_if_writable(ws, coordinate: str, fill: PatternFill) -> None:
    cell = ws[coordinate]
    if isinstance(cell, MergedCell):
        return
    cell.fill = copy(fill)


def _force_recalculate_on_open(wb) -> None:
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except AttributeError:
        return


def _num(value: float | None) -> str:
    if value is None:
        return "0"
    return f"{float(value):.8f}".rstrip("0").rstrip(".")


def _safe(value: str) -> str:
    text = str(value or "").strip() or "UNKNOWN"
    return re.sub(r'[<>:"/\\\\|?*\\s]+', "_", text).strip("_")
